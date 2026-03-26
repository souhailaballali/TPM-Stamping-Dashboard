"""

  TE Connectivity — Stamping Department
  Stamping CMMS — Full Version + Persistent Storage
  Bruderer Presses S-001 → S-006 + Peripherals

  INSTALLATION:
    pip install streamlit plotly pandas openpyxl numpy kaleido reportlab

  RUN:
    streamlit run app.py

"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, date, timedelta
import io
import os
import warnings
warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TE Connectivity — Stamping CMMS",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ──────────────────────────────────────────────────────────────────────────────
#  PERSISTENT STORAGE PATH
# ──────────────────────────────────────────────────────────────────────────────
PERSISTENT_CSV = "tpm_data_persistent.csv"
ARCHIVE_DIR    = "data_archive"        # auto-archive on every new import

# ──────────────────────────────────────────────────────────────────────────────
#  COLONNES SOURCE (noms exacts du fichier Hydra)
# ──────────────────────────────────────────────────────────────────────────────
COL_MACHINE   = "machine_id"
COL_STATUS    = "machine_status_name"
COL_DATE      = "plant_shift_date"
COL_MTTR      = "Sum of mttr_workcenter_numerator_seconds_quantity"
COL_MTBF      = "Sum of mtbf_numerator_seconds_quantity"
COL_PROD      = "hydra_bmk_production_status_name"
REQUIRED_COLS = [COL_MACHINE, COL_STATUS, COL_MTTR, COL_MTBF]

# ──────────────────────────────────────────────────────────────────────────────
#  COULEURS TE CONNECTIVITY
# ──────────────────────────────────────────────────────────────────────────────
TE_ORANGE  = "#E8650A"
TE_ORANGE2 = "#F0934A"
TE_ORANGE3 = "#F5B87A"
TE_ORANGE4 = "#FAD9B5"
TE_DARK    = "#C04D05"
TE_BLACK   = "#1C1C1C"
TE_NAVY    = "#1B2A4A"
TE_BROWN   = "#A07858"
TE_BG      = "#F7F4F0"
TE_WHITE   = "#FFFFFF"
TE_GREEN   = "#27AE60"
TE_RED     = "#C0392B"
TE_AMBER   = "#E67E22"

PALETTE = [TE_ORANGE, TE_NAVY, TE_RED, "#8E44AD", TE_GREEN, "#16A085", "#D4AC0D", TE_ORANGE2]

# ──────────────────────────────────────────────────────────────────────────────
#  CSS GLOBAL
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@300;400;500;600;700;800&family=Barlow+Condensed:wght@400;600;700;800&family=JetBrains+Mono:wght@300;400;500&display=swap');

/* ══════════════════════════════════════════════════
   ANTI-DARK MODE — Force Light Mode unconditionally
   ══════════════════════════════════════════════════ */
:root {{
    color-scheme: light only !important;
}}
html, body, .stApp, [data-testid="stAppViewContainer"],
[data-testid="stMain"], [data-testid="block-container"] {{
    background-color: {TE_BG} !important;
    font-family: 'Barlow', sans-serif;
    color-scheme: light !important;
}}
@media (prefers-color-scheme: dark) {{
    html, body, .stApp {{ background-color: {TE_BG} !important; }}
    [data-testid="stMain"], [data-testid="block-container"] {{
        background-color: {TE_BG} !important;
    }}
    input, textarea, select {{ background: white !important; color: #1C1C1C !important; }}
}}

#MainMenu, footer {{ visibility: hidden; }}
header[data-testid="stHeader"] {{ background: transparent !important; }}
.block-container {{ padding-top: 0 !important; max-width: 100% !important; }}

[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {TE_BLACK} 0%, #2A1A0A 100%) !important;
    border-right: 3px solid {TE_ORANGE} !important;
    min-width: 260px !important; max-width: 320px !important;
}}
[data-testid="stSidebar"] * {{
    color: #F0E8DF !important;
    font-family: 'Barlow', sans-serif !important;
}}
/* Restore the collapse button so it stays clickable and visible */
[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] {{
    background: transparent !important;
}}
[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button {{
    color: {TE_ORANGE} !important;
    background: transparent !important;
    opacity: 1 !important;
}}
[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button svg {{
    color: {TE_ORANGE} !important;
    fill: {TE_ORANGE} !important;
    stroke: {TE_ORANGE} !important;
}}
[data-testid="stSidebar"] > div:first-child {{ padding: 18px 16px !important; }}
[data-testid="stSidebar"] hr {{ border-color: #3D2A18 !important; }}
[data-testid="stSidebar"] h3 {{
    color: {TE_ORANGE} !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 11px !important; font-weight: 700 !important;
    letter-spacing: 3px !important; text-transform: uppercase !important;
}}
[data-testid="stSidebar"] [data-testid="stMultiSelect"] > div > div {{
    background-color: #2C1F14 !important;
    border: 1px solid {TE_ORANGE} !important;
    border-radius: 6px !important;
}}
[data-testid="stSidebar"] [data-testid="stMultiSelect"] span[data-baseweb="tag"] {{
    background-color: {TE_ORANGE} !important; border-radius: 4px !important;
}}
[data-testid="stSidebar"] [data-testid="stMultiSelect"] span[data-baseweb="tag"] span {{
    color: white !important; font-weight: 700 !important; font-size: 11px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stFileUploadDropzone"] {{
    background: #FFF8F2 !important;
    border: 1.5px dashed {TE_ORANGE} !important;
    border-radius: 10px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stFileUploadDropzone"] * {{
    color: #2e1808 !important; opacity: 1 !important;
}}
section[data-testid="stSidebar"] div[data-testid="stFileUploadDropzone"] p,
section[data-testid="stSidebar"] div[data-testid="stFileUploadDropzone"] span,
section[data-testid="stSidebar"] div[data-testid="stFileUploadDropzone"] small {{
    color: #2e1808 !important; opacity: 1 !important; font-weight: 700 !important;
}}

.te-header {{
    background: linear-gradient(135deg, {TE_BLACK} 0%, #2A1A0A 60%, #3D2508 100%);
    border-radius: 14px; padding: 26px 36px; margin-bottom: 20px;
    border-left: 6px solid {TE_ORANGE};
    box-shadow: 0 8px 32px rgba(232,101,10,0.18);
    position: relative; overflow: hidden;
}}
.te-header::after {{
    content: ''; position: absolute; top: -40px; right: -40px;
    width: 180px; height: 180px;
    background: radial-gradient(circle, rgba(232,101,10,0.15) 0%, transparent 70%);
    border-radius: 50%;
}}
.te-header-tag {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 9px; font-weight: 700; letter-spacing: 3px;
    text-transform: uppercase; color: {TE_ORANGE}; margin-bottom: 6px;
    display: flex; align-items: center; gap: 8px;
}}
.te-header-tag::before {{
    content: ''; width: 20px; height: 2px;
    background: {TE_ORANGE}; border-radius: 1px;
}}
.te-header-title {{
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 38px; font-weight: 800; color: {TE_WHITE};
    line-height: 1.0; margin-bottom: 4px; text-transform: uppercase;
    letter-spacing: 0.5px;
}}
.te-header-title span {{ color: {TE_ORANGE}; }}
.te-header-sub {{
    font-size: 13px; color: {TE_BROWN}; font-weight: 400; margin-bottom: 10px;
}}
.te-header-badge {{
    display: inline-flex; align-items: center; gap: 6px;
    background: rgba(232,101,10,0.15); border: 1px solid rgba(232,101,10,0.4);
    color: {TE_ORANGE2}; font-size: 11px; font-weight: 600;
    letter-spacing: 1px; text-transform: uppercase;
    border-radius: 20px; padding: 4px 14px;
}}
.te-header-right {{
    display: flex; flex-direction: column; align-items: flex-end; gap: 6px;
}}
.te-live {{
    display: flex; align-items: center; gap: 6px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; letter-spacing: 1px; color: {TE_GREEN};
    background: rgba(39,174,96,0.12); border: 1px solid rgba(39,174,96,0.3);
    border-radius: 20px; padding: 5px 14px;
}}
.te-live-dot {{
    width: 7px; height: 7px; border-radius: 50%; background: {TE_GREEN};
    animation: blink 2s infinite;
}}
@keyframes blink {{
    0%,100% {{ opacity:1; transform:scale(1); }}
    50%      {{ opacity:0.3; transform:scale(1.5); }}
}}

.te-statusbar {{
    display: flex; align-items: center; gap: 18px;
    background: {TE_WHITE}; border: 1px solid #EDE0D4; border-radius: 10px;
    padding: 10px 20px; margin-bottom: 20px;
    font-size: 13px; color: #7A6050; flex-wrap: wrap;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
}}
.te-statusbar strong {{ color: {TE_BLACK}; font-weight: 600; }}
.te-sep {{ width: 1px; height: 16px; background: #E0D0C0; flex-shrink: 0; }}
.te-statusbar-item {{ display: flex; align-items: center; gap: 5px; }}
.te-dot-green {{
    width: 8px; height: 8px; background: {TE_GREEN}; border-radius: 50%;
    box-shadow: 0 0 6px rgba(39,174,96,0.5); flex-shrink: 0;
}}

.kpi-card {{
    background: {TE_BLACK}; border-radius: 12px; padding: 22px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.30); border: 1px solid #2E2E2E;
    position: relative; overflow: hidden;
    transition: transform 0.2s, box-shadow 0.2s;
}}
.kpi-card:hover {{
    transform: translateY(-3px);
    box-shadow: 0 8px 32px rgba(232,101,10,0.25);
}}
.kpi-card::before {{
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 4px;
    background: linear-gradient(90deg, {TE_ORANGE}, {TE_ORANGE3});
    border-radius: 12px 12px 0 0;
}}
.kpi-label {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 9px; font-weight: 700; letter-spacing: 2.5px;
    text-transform: uppercase; color: {TE_ORANGE}; margin-bottom: 4px;
}}
.kpi-value {{
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 38px; font-weight: 700; color: {TE_WHITE};
    line-height: 1; margin-bottom: 4px; letter-spacing: -0.5px;
}}
.kpi-divider {{
    width: 28px; height: 3px;
    background: linear-gradient(90deg, {TE_ORANGE}, {TE_ORANGE3});
    border-radius: 2px; margin: 6px 0;
}}
.kpi-unit {{ font-size: 11px; color: {TE_BROWN}; font-weight: 500; }}

.te-section {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 9px; font-weight: 700; letter-spacing: 3px;
    text-transform: uppercase; color: {TE_ORANGE};
    margin: 24px 0 12px 0;
    display: flex; align-items: center; gap: 10px;
}}
.te-section::before {{
    content: ''; width: 20px; height: 2px;
    background: {TE_ORANGE}; border-radius: 1px;
}}
.te-section::after {{
    content: ''; flex: 1; height: 1px;
    background: linear-gradient(90deg, #F0D0B0, transparent);
}}

.chart-card {{
    background: {TE_WHITE}; border-radius: 12px;
    padding: 18px 18px 8px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    border: 1px solid #EDE0D4; margin-bottom: 18px;
}}
.chart-header {{
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 12px; padding-bottom: 10px;
    border-bottom: 1px solid #F0E4D8;
}}
.chart-dot {{
    width: 10px; height: 10px; background: {TE_ORANGE};
    border-radius: 50%; flex-shrink: 0;
}}
.chart-title {{
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 15px; font-weight: 700; color: {TE_BLACK};
    letter-spacing: 0.5px; text-transform: uppercase;
}}

.quad-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 12px; }}
.quad {{ padding: 11px 14px; border-radius: 8px; border: 1px solid; }}
.quad h5 {{ font-family:'Barlow Condensed',sans-serif; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:3px; }}
.quad p  {{ font-size:10px; line-height:1.5; margin:0; }}
.q-good   {{ background:#eafaf1; border-color:#a9dfbf; }} .q-good h5   {{ color:#1e8449; }} .q-good p   {{ color:#145a32; }}
.q-watch  {{ background:#eaf2ff; border-color:#aed6f1; }} .q-watch h5  {{ color:#1a5276; }} .q-watch p  {{ color:#1a3d6d; }}
.q-warn   {{ background:#fef9e7; border-color:#f9e79f; }} .q-warn h5   {{ color:#d68910; }} .q-warn p   {{ color:#7d6608; }}
.q-crit   {{ background:#fdf2f2; border-color:#e8a0a0; }} .q-crit h5   {{ color:#c0392b; }} .q-crit p   {{ color:#7b241c; }}

.te-insight {{
    background: #FEF0E1; border: 1px solid #FAC98A;
    border-left: 4px solid {TE_ORANGE}; border-radius: 8px;
    padding: 12px 16px; font-size: 12px; color: #4A3020; line-height: 1.65;
    margin-top: 12px;
}}
.te-insight strong {{ color: {TE_BLACK}; }}
.te-insight-crit {{
    background: #fdf2f2; border: 1px solid #e8a0a0;
    border-left: 4px solid {TE_RED}; border-radius: 8px;
    padding: 12px 16px; font-size: 12px; color: #4A3020; line-height: 1.65;
    margin-top: 12px;
}}
.te-insight-ok {{
    background: #eafaf1; border: 1px solid #a9dfbf;
    border-left: 4px solid {TE_GREEN}; border-radius: 8px;
    padding: 12px 16px; font-size: 12px; color: #4A3020; line-height: 1.65;
    margin-top: 12px;
}}

/* History page — dark TE style (matches Dashboard) */
.hist-header {{
    background: linear-gradient(135deg, {TE_BLACK} 0%, #2A1A0A 60%, #3D2508 100%);
    border-radius: 14px; padding: 26px 36px; margin-bottom: 20px;
    border-left: 6px solid {TE_ORANGE};
    box-shadow: 0 8px 32px rgba(232,101,10,0.18);
    position: relative; overflow: hidden;
}}
.hist-kpi-card {{
    background: {TE_BLACK}; border-radius: 12px; padding: 22px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.30); border: 1px solid #2E2E2E;
    position: relative; overflow: hidden;
    transition: transform 0.2s, box-shadow 0.2s;
}}
.hist-kpi-card::before {{
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 4px;
    background: linear-gradient(90deg, {TE_ORANGE}, {TE_ORANGE3});
    border-radius: 12px 12px 0 0;
}}
.hist-kpi-label {{
    font-family: 'JetBrains Mono', monospace; font-size: 9px;
    font-weight: 700; letter-spacing: 2.5px; text-transform: uppercase;
    color: {TE_ORANGE}; margin-bottom: 4px;
}}
.hist-kpi-value {{
    font-family: 'Barlow Condensed', sans-serif; font-size: 34px;
    font-weight: 700; color: {TE_WHITE}; line-height: 1;
}}
.hist-kpi-unit {{
    font-size: 11px; color: {TE_BROWN}; font-weight: 500; margin-top: 4px;
}}

.stDownloadButton > button {{
    background: linear-gradient(135deg, {TE_ORANGE}, {TE_DARK}) !important;
    color: white !important; border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; padding: 8px 16px !important;
    box-shadow: 0 4px 14px rgba(232,101,10,0.3) !important;
    width: 100% !important; font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 13px !important; letter-spacing: 0.5px !important; text-transform: uppercase !important;
}}
.stDownloadButton > button:hover {{
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(232,101,10,0.45) !important;
}}
.stButton > button {{
    background: {TE_NAVY} !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 13px !important; font-weight: 700 !important;
    text-transform: uppercase !important; letter-spacing: 0.5px !important;
    padding: 9px 20px !important;
}}

[data-testid="stExpander"] {{
    background: {TE_WHITE} !important;
    border: 1px solid #EDE0D4 !important; border-radius: 12px !important;
}}
[data-testid="stExpander"] summary {{
    font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 700 !important; color: {TE_BLACK} !important;
    font-size: 14px !important; text-transform: uppercase !important;
    letter-spacing: 0.5px !important;
    background: #FFFAF6 !important; padding: 14px 20px !important;
}}

[data-testid="stTabs"] [data-baseweb="tab"] {{
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 14px !important; font-weight: 700 !important;
    letter-spacing: 1px !important; text-transform: uppercase !important;
}}
[data-testid="stTabs"] [aria-selected="true"] {{
    color: {TE_ORANGE} !important;
    border-bottom: 3px solid {TE_ORANGE} !important;
}}

[data-testid="stDataFrame"] {{
    border: 1px solid #EDE0D4 !important;
    border-radius: 10px !important; overflow: hidden !important;
}}

::-webkit-scrollbar {{ width: 5px; height: 5px; }}
::-webkit-scrollbar-track {{ background: #F0EAE3; }}
::-webkit-scrollbar-thumb {{ background: {TE_ORANGE3}; border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: {TE_ORANGE}; }}

/* ── Sidebar toggle button — inside sidebar ── */
[data-testid="stSidebar"] [data-testid="stButton"]:has(button[kind="secondary"]) button,
[data-testid="stSidebar"] .stButton button {{
    background: linear-gradient(135deg, {TE_ORANGE} 0%, {TE_DARK} 100%) !important;
    color: white !important;
    border: 2px solid {TE_ORANGE3} !important;
    border-radius: 10px !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 14px !important;
    font-weight: 800 !important;
    letter-spacing: 2px !important;
    text-transform: uppercase !important;
    padding: 12px 20px !important;
    width: 100% !important;
    box-shadow: 0 4px 18px rgba(232,101,10,0.45), 0 0 0 1px rgba(232,101,10,0.2) !important;
    transition: all 0.2s ease !important;
    cursor: pointer !important;
}}
[data-testid="stSidebar"] .stButton button:hover {{
    background: linear-gradient(135deg, {TE_ORANGE3} 0%, {TE_ORANGE} 100%) !important;
    box-shadow: 0 6px 28px rgba(232,101,10,0.65), 0 0 0 2px rgba(232,101,10,0.4) !important;
    transform: translateY(-1px) !important;
}}

/* ── Sidebar toggle — native buttons, minimal styling only ── */
[data-testid="collapsedControl"] button svg {{
    color: {TE_ORANGE} !important; fill: {TE_ORANGE} !important;
}}
[data-testid="stSidebarCollapseButton"] button svg {{
    color: {TE_ORANGE} !important; fill: {TE_ORANGE} !important;
}}
/* ── Custom HIDE button inside sidebar ── */
[data-testid="stSidebar"] .te-hide-btn button {{
    background: linear-gradient(135deg, {TE_ORANGE} 0%, {TE_DARK} 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 13px !important;
    font-weight: 800 !important;
    letter-spacing: 2px !important;
    text-transform: uppercase !important;
    width: 100% !important;
    padding: 10px 16px !important;
    box-shadow: 0 3px 12px rgba(232,101,10,0.4) !important;
    transition: all 0.2s ease !important;
    cursor: pointer !important;
}}
[data-testid="stSidebar"] .te-hide-btn button:hover {{
    background: linear-gradient(135deg, {TE_ORANGE3} 0%, {TE_ORANGE} 100%) !important;
    box-shadow: 0 5px 20px rgba(232,101,10,0.65) !important;
    transform: translateY(-1px) !important;
}}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
#  LISTES DÉROULANTES — Editable Table
# ──────────────────────────────────────────────────────────────────────────────
SHIFTS = ["", "A (6-14h)", "B (14-22h)", "C (22-6h)"]

KEY_FAILURES = [
    "", "Laser monitoring", "Laser temp too high", "Dirt detected on the part",
    "part pos variation in front of the cam", "Hydra count issue", "Force monitoring",
    "Feeder issue", "Label print issue", "Camera loop variation", "Strip welding issue",
    "Reeling errors", "Dereeling errors", "air pressure error", "oil pressure error",
    "Raziol lubrication issue", "Bond welding detection issue", "camera strip driving",
    "Electrical power issue", "Bruderer system issue", "Extraction issue",
    "Cooling system issue", "Compressed air issue", "Camera setting issue",
    "Camera hardware issue", "Laser communication issue", "Laser water level",
    "Tooling detection", "Safety detection", "Blowing system error", "Hydraulic issue",
    "Equipement changeover", "Laser HW issue", "Laser internal cooling", "Machine guarding",
]

_QUAL_COLS_DEFAULTS = [
    ("User ID",           "TE"),
    ("Shift",             ""),
    ("Key Failure",       ""),
    ("Issue Description", ""),
    ("Action Taken",      ""),
    ("Spare Part Ref",    ""),
    ("Qty",               0),
    ("Unit Price (€)",    0.0),
    ("Total Part Cost",   0.0),
]

# ──────────────────────────────────────────────────────────────────────────────
#  PERSISTENT STORAGE FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

def load_persistent() -> pd.DataFrame:
    if os.path.exists(PERSISTENT_CSV):
        try:
            df_p = pd.read_csv(PERSISTENT_CSV, low_memory=False)
            df_p.columns = [str(c).strip() for c in df_p.columns]
            return df_p
        except Exception as e:
            st.warning(f" Could not read persistent file: {e}")
    return pd.DataFrame()


def save_persistent(df_to_save: pd.DataFrame):
    """Save main persistent CSV (overwrite). No backup copy on save."""
    try:
        df_to_save.to_csv(PERSISTENT_CSV, index=False, encoding="utf-8")
    except Exception as e:
        st.error(f" Could not save to disk: {e}")


def archive_import(df_source: pd.DataFrame):
    """
    On every new file import, save a copy to data_archive/ named:
    'DATA [DateMin] - [DateMax].csv'
    using the min/max of plant_shift_date column.
    """
    try:
        os.makedirs(ARCHIVE_DIR, exist_ok=True)
        # Extract date range
        if COL_DATE in df_source.columns:
            _dates = pd.to_datetime(df_source[COL_DATE], errors="coerce").dropna()
            if len(_dates):
                d_min = _dates.min().strftime("%d-%m-%Y")
                d_max = _dates.max().strftime("%d-%m-%Y")
            else:
                d_min = d_max = datetime.now().strftime("%d-%m-%Y")
        else:
            d_min = d_max = datetime.now().strftime("%d-%m-%Y")
        fname = f"DATA {d_min} - {d_max}.csv"
        fpath = os.path.join(ARCHIVE_DIR, fname)
        df_source.to_csv(fpath, index=False, encoding="utf-8")
        return fname
    except Exception:
        return None


def list_archive() -> list:
    """
    Return list of (display_label, filepath) from data_archive/, newest first.
    Label = filename without .csv extension.
    """
    if not os.path.exists(ARCHIVE_DIR):
        return []
    files = sorted(
        [f for f in os.listdir(ARCHIVE_DIR) if f.endswith(".csv")],
        reverse=True
    )
    return [(f.replace(".csv", ""), os.path.join(ARCHIVE_DIR, f)) for f in files]


def load_archive(filepath: str) -> pd.DataFrame:
    try:
        df_b = pd.read_csv(filepath, low_memory=False)
        df_b.columns = [str(c).strip() for c in df_b.columns]
        return df_b
    except Exception as e:
        st.error(f"Cannot load archive file: {e}")
        return pd.DataFrame()


def merge_qualifications(df_new: pd.DataFrame,
                          df_persist: pd.DataFrame) -> pd.DataFrame:
    if df_persist.empty:
        return df_new
    QUAL_COLS = [c for c, _ in _QUAL_COLS_DEFAULTS]
    KEY_COLS  = [c for c in [COL_MACHINE, COL_DATE, COL_STATUS]
                 if c in df_new.columns and c in df_persist.columns]
    if not KEY_COLS:
        return df_new
    _has_data = df_persist[
        [c for c in QUAL_COLS if c in df_persist.columns]
    ].apply(
        lambda r: any(str(v).strip() not in ("", "nan", "None", "0", "0.0") for v in r),
        axis=1
    )
    df_qual_only = df_persist[_has_data].copy()
    if df_qual_only.empty:
        return df_new
    for _c in KEY_COLS:
        df_qual_only[_c] = df_qual_only[_c].astype(str).str.strip()
        df_new[_c]       = df_new[_c].astype(str).str.strip()
    _merge_cols = KEY_COLS + [c for c in QUAL_COLS if c in df_qual_only.columns]
    df_merged = df_new.merge(
        df_qual_only[_merge_cols].drop_duplicates(subset=KEY_COLS),
        on=KEY_COLS, how="left", suffixes=("", "_PERSIST"))
    for _c in QUAL_COLS:
        _pc = f"{_c}_PERSIST"
        if _pc in df_merged.columns:
            _default = str(_QUAL_COLS_DEFAULTS[
                [x[0] for x in _QUAL_COLS_DEFAULTS].index(_c)
            ][1]) if _c in [x[0] for x in _QUAL_COLS_DEFAULTS] else ""
            _blank = df_merged[_c].astype(str).str.strip().isin(
                ["", "nan", "None", str(_default)])
            df_merged.loc[_blank, _c] = df_merged.loc[_blank, _pc]
            df_merged.drop(columns=[_pc], inplace=True)
    return df_merged

# ──────────────────────────────────────────────────────────────────────────────
#  DATA LOADING HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def load_data(f) -> pd.DataFrame:
    """
    Ultra-robust file reader.
    - Accepts UploadedFile or a file-path string.
    - Bytes are cached in session_state to prevent double-read errors.
    - Tries multiple CSV strategies (sep=None auto-detect, then ;, then ,).
    - Never raises: returns empty DataFrame on any failure.
    """
    # ── File path (string) — for archive/library loads ─────────────
    if isinstance(f, str):
        try:
            if f.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(f)
            else:
                # Try auto-detect first, then fallbacks
                try:
                    df = pd.read_csv(f, sep=None, engine="python",
                                     on_bad_lines="skip", low_memory=False)
                except Exception:
                    try:
                        df = pd.read_csv(f, sep=";", on_bad_lines="skip", low_memory=False)
                    except Exception:
                        df = pd.read_csv(f, sep=",", on_bad_lines="skip", low_memory=False)
            df.columns = [str(c).strip() for c in df.columns]
            return df if not df.empty else pd.DataFrame()
        except Exception:
            return pd.DataFrame()

    # ── UploadedFile — read bytes ONCE, cache to survive reruns ────
    cache_key = f"_raw_{f.name}_{f.size}"
    if cache_key not in st.session_state:
        try:
            raw = f.read()
            st.session_state[cache_key] = raw
        except Exception:
            return pd.DataFrame()

    raw = st.session_state.get(cache_key, b"")
    if not raw:
        return pd.DataFrame()

    name = f.name.lower()
    try:
        if name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(raw))
            df.columns = [str(c).strip() for c in df.columns]
            return df if len(df.columns) > 0 else pd.DataFrame()

        # CSV: try sep=None (auto), then ;, then ,
        for _sep, _eng in [(None, "python"), (";", "c"), (",", "c")]:
            try:
                _kwargs = dict(on_bad_lines="skip", low_memory=False)
                if _sep is None:
                    _kwargs["sep"]    = None
                    _kwargs["engine"] = "python"
                else:
                    _kwargs["sep"] = _sep
                df = pd.read_csv(io.BytesIO(raw), **_kwargs)
                df.columns = [str(c).strip() for c in df.columns]
                # Accept if we got at least 2 columns (not a single-col parse failure)
                if len(df.columns) >= 2:
                    return df
            except Exception:
                continue
        # Last resort: try with utf-8-sig encoding (BOM)
        try:
            df = pd.read_csv(io.BytesIO(raw), sep=None, engine="python",
                             encoding="utf-8-sig", on_bad_lines="skip")
            df.columns = [str(c).strip() for c in df.columns]
            if len(df.columns) >= 2:
                return df
        except Exception:
            pass
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def check_missing(df: pd.DataFrame) -> list:
    return [c for c in REQUIRED_COLS if c not in df.columns]


def fmt(val, decimals=2):
    if pd.isna(val): return "—"
    if val >= 1_000_000: return f"{val/1_000_000:.{decimals}f}M"
    if val >= 1000:      return f"{val/1000:.{decimals}f}k"
    return f"{val:,.{decimals}f}"


def sec_to_h(s):
    return round(float(s) / 3600.0, 4) if not pd.isna(s) else 0.0


def dl_png(fig, filename, label="Download PNG"):
    try:
        img = fig.to_image(format="png", width=1400, height=680, scale=2)
        st.download_button(label=label, data=img, file_name=filename,
                           mime="image/png", use_container_width=True)
    except Exception:
        st.caption("_`pip install kaleido` to enable PNG export_")


def export_excel(df: pd.DataFrame, kpi: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    try:
        wb = Workbook()
        def hdr_style(ws, row, col, value, bg="E8650A", fg="FFFFFF"):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color=fg, name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            return cell
        def title_style(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color="1B2A4A", name="Arial", size=15)
            return cell
        def auto_width(ws):
            for col in ws.columns:
                max_w = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 50)
        ws1 = wb.active
        ws1.title = "KPI Summary"
        ws1.row_dimensions[1].height = 22
        title_style(ws1, 1, 1, "TE Connectivity — Stamping CMMS Report")
        ws1.cell(2, 1, f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        hdr_style(ws1, 4, 1, "Indicator")
        hdr_style(ws1, 4, 2, "Value")
        kpi_rows = [
            ("Availability (%)",       f"{kpi['dispo']:.2f}%"),
            ("Mean MTTR / stop (h)",   f"{kpi['mttr_mean_h']:.4f} h"),
            ("Mean MTBF (h)",          f"{kpi['mtbf_mean_h']:.4f} h"),
            ("Total Failures",         str(kpi['nb_arrets'])),
            ("Cumulative Downtime (h)",f"{kpi['mttr_total_h']:.2f} h"),
            ("Cumulative Uptime (h)",  f"{kpi['mtbf_total_h']:.2f} h"),
            ("Total events analyzed",  str(kpi['nb_rows'])),
        ]
        for i, (k, v) in enumerate(kpi_rows, start=5):
            ws1.cell(i, 1, k)
            ws1.cell(i, 2, v)
        auto_width(ws1)
        if "by_machine" in kpi and not kpi["by_machine"].empty:
            ws2 = wb.create_sheet("By Machine")
            bm  = kpi["by_machine"].copy()
            # ma has columns: machine_id, mean_mttr_h, mean_mtbf_h, nb_failures, nb_events, dispo
            bm  = bm[[COL_MACHINE, "mean_mttr_h", "mean_mtbf_h", "nb_failures", "nb_events", "dispo"]].copy()
            bm.columns = ["Machine","Mean MTTR (h)","Mean MTBF (h)","Failures","Events","Availability (%)"]
            for ci, col_name in enumerate(bm.columns, start=1):
                hdr_style(ws2, 1, ci, col_name, bg="1B2A4A")
            for ri, row_vals in enumerate(bm.itertuples(index=False), start=2):
                for ci, v in enumerate(row_vals, start=1):
                    ws2.cell(ri, ci, round(v, 4) if isinstance(v, float) else (int(v) if hasattr(v, "item") else v))
            auto_width(ws2)
        if "pareto" in kpi and not kpi["pareto"].empty:
            ws3 = wb.create_sheet("Pareto Downtime")
            par = kpi["pareto"]
            for ci, col_name in enumerate(par.columns, start=1):
                hdr_style(ws3, 1, ci, col_name)
            for ri, row_vals in enumerate(par.itertuples(index=False), start=2):
                for ci, v in enumerate(row_vals, start=1):
                    ws3.cell(ri, ci, round(v, 4) if isinstance(v, float) else v)
            auto_width(ws3)
        ws4 = wb.create_sheet("Filtered Data")
        exp_cols = [c for c in [COL_MACHINE, COL_DATE, COL_STATUS,
                                "mttr_h", "mtbf_h"] if c in df.columns]
        for ci, col_name in enumerate(exp_cols, start=1):
            hdr_style(ws4, 1, ci, col_name)
        for ri, row_vals in enumerate(df[exp_cols].itertuples(index=False), start=2):
            for ci, v in enumerate(row_vals, start=1):
                ws4.cell(ri, ci, str(v) if not isinstance(v, (int, float)) else v)
        auto_width(ws4)
        wb.save(buf)
    except Exception as e:
        st.error(f"Excel export error: {e}")
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
#  PLOTLY BASE LAYOUT
# ──────────────────────────────────────────────────────────────────────────────
PL = dict(
    plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
    font=dict(family="Barlow, sans-serif", color="#4A3020", size=11),
    margin=dict(l=20, r=20, t=40, b=20),
    xaxis=dict(gridcolor="#F0E8E0", showgrid=True, zeroline=False,
               linecolor="#EDE0D4", tickfont=dict(size=10, color="#9A7A60"),
               color="#4A3020"),
    yaxis=dict(gridcolor="#F0E8E0", showgrid=True, zeroline=False,
               linecolor="#EDE0D4", tickfont=dict(size=10, color="#9A7A60"),
               color="#4A3020"),
    legend=dict(bgcolor="#FFFFFF", bordercolor="#EDE0D4", borderwidth=1,
                font=dict(size=11, color="#4A3020")),
    hoverlabel=dict(bgcolor=TE_BLACK, bordercolor=TE_BLACK,
                    font=dict(color="white", family="JetBrains Mono", size=11)),
    template="plotly_white",
)
PCONF = dict(displayModeBar=False, responsive=True)

def apply(fig, **kw):
    fig.update_layout(**{**PL, **kw})
    return fig

def _hex_to_rgba(hex_color: str, alpha: float = 0.1) -> str:
    h = hex_color.lstrip("#")[:6]
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"




# ──────────────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:

    st.markdown(f"""
    <div style="background:rgba(232,101,10,0.12);border:1px solid rgba(232,101,10,0.35);
                border-radius:10px;padding:12px 14px;margin-bottom:16px">
        <div style="font-family:'Barlow Condensed',sans-serif;font-size:20px;
                    font-weight:800;letter-spacing:1.5px;color:{TE_ORANGE}">
            ≡ TE CONNECTIVITY
        </div>
        <div style="font-family:'JetBrains Mono',monospace;font-size:7px;
                    letter-spacing:2px;color:rgba(255,255,255,0.4);margin-top:4px">
            STAMPING CMMS · TANGIER
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Navigation ──
    st.markdown(f'<p style="font-size:9px;font-weight:700;letter-spacing:3px;'
                f'text-transform:uppercase;color:{TE_ORANGE};margin-bottom:6px">'
                f'NAVIGATION</p>', unsafe_allow_html=True)
    nav_choice = st.radio(
        "", options=["Dashboard", "History"],
        index=0, key="nav_radio", label_visibility="collapsed"
    )
    st.markdown("---")

    # ── Import Data ──
    st.markdown(f'<p style="font-size:9px;font-weight:700;letter-spacing:3px;'
                f'text-transform:uppercase;color:{TE_ORANGE};margin-bottom:6px">'
                f'IMPORT DATA</p>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "", type=["csv", "xlsx", "xls"],
        key="sidebar_uploader", label_visibility="collapsed"
    )

    # ── Auto-load indicator ──
    if os.path.exists(PERSISTENT_CSV):
        _sz = os.path.getsize(PERSISTENT_CSV)
        _mt = datetime.fromtimestamp(os.path.getmtime(PERSISTENT_CSV))
        st.markdown(f"""
        <div style="background:rgba(39,174,96,0.18);border:1px solid rgba(39,174,96,0.5);
                    border-radius:8px;padding:9px 12px;margin:10px 0 2px">
          <div style="font-size:11px;font-weight:700;color:#4AE080;margin-bottom:2px">
            ✅ Data loaded from last session
          </div>
          <div style="font-size:9px;color:rgba(240,232,223,0.6);
                      font-family:'JetBrains Mono',monospace;letter-spacing:0.5px">
            {_sz//1024} KB · {_mt.strftime('%d/%m/%Y %H:%M')}
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    if uploaded is not None:
        st.markdown(f'<p style="font-size:9px;font-weight:700;letter-spacing:3px;'
                    f'text-transform:uppercase;color:{TE_ORANGE};margin-bottom:6px">'
                    f'FILTERS</p>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    #  📌 DATA LIBRARY
    # ══════════════════════════════════════════════════════════════
    _archive_files = list_archive()
    if _archive_files:
        st.markdown("---")
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">
          <span style="font-size:14px">📌</span>
          <span style="font-size:9px;font-weight:700;letter-spacing:3px;
                       text-transform:uppercase;color:{TE_ORANGE}">DATA LIBRARY</span>
          <span style="font-size:9px;color:rgba(240,232,223,0.4);
                       font-family:'JetBrains Mono',monospace">
            ({len(_archive_files)} session{"s" if len(_archive_files)>1 else ""})
          </span>
        </div>
        """, unsafe_allow_html=True)

        _arch_labels = [lbl for lbl, _ in _archive_files]
        _arch_paths  = {lbl: fp for lbl, fp in _archive_files}

        # ── Highlight active session ──
        _active_lbl = ""
        if st.session_state.get("library_active"):
            _af = os.path.basename(st.session_state.library_active).replace(".csv","")
            if _af in _arch_labels:
                _active_lbl = _af
        if _active_lbl:
            st.markdown(
                f'<div style="font-size:9px;color:{TE_ORANGE};font-family:'
                f"'JetBrains Mono',monospace;margin-bottom:4px\">"
                f"▶ Active: {_active_lbl}</div>",
                unsafe_allow_html=True)

        _sel_session = st.selectbox(
            "📁 Select Session",
            options=["— choose —"] + _arch_labels,
            key="lib_selectbox",
            label_visibility="visible"
        )
        _c1, _c2 = st.columns([3, 1])
        with _c1:
            if st.button("▶ Load", key="btn_load_library",
                         use_container_width=True,
                         disabled=(_sel_session == "— choose —")):
                _lib_df = load_archive(_arch_paths[_sel_session])
                if not _lib_df.empty:
                    st.session_state.library_df     = _lib_df
                    st.session_state.library_active = _arch_paths[_sel_session]
                    st.session_state.edited_df      = None
                    st.session_state.last_file      = f"__library__{_arch_paths[_sel_session]}"
                    st.rerun()
        with _c2:
            if st.button("🗑", key="btn_del_archive",
                         use_container_width=True,
                         disabled=(_sel_session == "— choose —"),
                         help="Delete this archive file"):
                try:
                    _del_path = _arch_paths[_sel_session]
                    # Windows-safe: overwrite with empty then remove
                    open(_del_path, "w").close()
                    os.remove(_del_path)
                    if st.session_state.get("library_active") == _del_path:
                        st.session_state.library_df     = None
                        st.session_state.library_active = None
                    st.success(f"🗑 Deleted: {_sel_session}")
                    st.rerun()
                except Exception as _de:
                    st.error(f"Cannot delete: {_de}")

    if st.session_state.get("_archive_saved"):
        _saved_name = st.session_state.pop("_archive_saved")
        st.success(f"✅ Archived: {_saved_name}")


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: HISTORY
# ══════════════════════════════════════════════════════════════════════════════
if nav_choice == "History":

    st.markdown(f"""
    <div class="hist-header">
      <div style="display:flex;justify-content:space-between;align-items:center">
        <div>
          <div style="font-family:'JetBrains Mono',monospace;font-size:9px;
                      font-weight:700;letter-spacing:3px;text-transform:uppercase;
                      color:{TE_ORANGE};margin-bottom:6px">
            Stamping Department · Bruderer Presses
          </div>
          <div style="font-family:'Barlow Condensed',sans-serif;font-size:38px;
                      font-weight:800;color:{TE_WHITE};text-transform:uppercase;
                      letter-spacing:0.5px;margin-bottom:4px;line-height:1.0">
            QUALIFIED STOPS <span style="color:{TE_ORANGE}">HISTORY</span>
          </div>
          <div style="font-size:12px;color:{TE_BROWN};margin-top:4px">
            Persistent data loaded from <code style="color:{TE_ORANGE2};background:rgba(232,101,10,0.12);padding:2px 6px;border-radius:4px">{PERSISTENT_CSV}</code>
          </div>
        </div>
        <div style="text-align:right">
          <div style="font-family:'JetBrains Mono',monospace;font-size:10px;
                      color:rgba(255,255,255,0.35)">TANGIER · PLANT 1310</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    df_hist = load_persistent()

    if df_hist.empty:
        st.markdown(f"""
        <div style="background:{TE_WHITE};border:2px dashed #F0C8A0;border-radius:14px;
                    padding:48px;text-align:center;margin:32px auto;max-width:480px">
          <div style="font-size:40px;margin-bottom:12px"></div>
          <div style="font-family:'Barlow Condensed',sans-serif;font-size:22px;
                      font-weight:800;color:{TE_BLACK};text-transform:uppercase;
                      margin-bottom:8px">No History Yet</div>
          <div style="font-size:13px;color:#9A7A60;line-height:1.7">
            No persistent data found.<br>
            Import a Hydra file, qualify some stops<br>
            and click <strong> Save Changes</strong> to populate this page.
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    # Build columns available
    QUAL_CHECK = ["Shift", "Key Failure", "Issue Description", "Action Taken", "Spare Part Ref"]

    def _is_qualified_hist(r):
        return any(str(r.get(c, "")).strip() not in ("", "None", "nan")
                   for c in QUAL_CHECK if c in r.index)

    _stop_mask = pd.to_numeric(df_hist.get("mttr_h", pd.Series(dtype=float)),
                               errors="coerce").fillna(0) > 0
    df_stops_hist = df_hist[_stop_mask].copy() if "mttr_h" in df_hist.columns else df_hist.copy()
    _qual_mask = df_stops_hist.apply(_is_qualified_hist, axis=1) if not df_stops_hist.empty else pd.Series(dtype=bool)
    df_qualified_hist = df_stops_hist[_qual_mask].copy() if len(_qual_mask) > 0 else pd.DataFrame()

    # ── Summary KPIs ──
    _tot_rows  = len(df_hist)
    _tot_stops = len(df_stops_hist)
    _tot_qual  = len(df_qualified_hist)
    _tot_cost  = float(
        pd.to_numeric(df_hist.get("Total Part Cost", pd.Series([0.0])),
                      errors="coerce").fillna(0).sum()
    ) if "Total Part Cost" in df_hist.columns else 0.0
    _pct_qual  = (_tot_qual / _tot_stops * 100) if _tot_stops > 0 else 0.0

    hc1, hc2, hc3, hc4 = st.columns(4)
    for _col, _lbl, _val, _unit in [
        (hc1, "TOTAL EVENTS",    str(_tot_rows),                        "in persistent file"),
        (hc2, "TOTAL STOPS",     str(_tot_stops),                       "with MTTR > 0"),
        (hc3, "QUALIFIED STOPS", f"{_tot_qual} ({_pct_qual:.0f}%)",     "shift + key failure"),
        (hc4, "SPARE PARTS COST",f"€ {_tot_cost:,.2f}",                "total recorded"),
    ]:
        with _col:
            st.markdown(f"""
            <div class="hist-kpi-card">
              <div class="hist-kpi-label">{_lbl}</div>
              <div class="hist-kpi-value">{_val}</div>
              <div class="hist-kpi-unit">{_unit}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if df_qualified_hist.empty:
        st.info("No qualified stops in the persistent file yet. Fill in Shift/Key Failure in the Dashboard and Save Changes.")
        st.stop()

    # ── FILTER HISTORY ──
    st.markdown(f'<div class="te-section">FILTER HISTORY</div>', unsafe_allow_html=True)
    _hf1, _hf2, _hf3 = st.columns(3)

    with _hf1:
        _hmachines = ["All"] + sorted(
            df_qualified_hist[COL_MACHINE].dropna().unique().tolist()
        ) if COL_MACHINE in df_qualified_hist.columns else ["All"]
        _hfm = st.selectbox(" Machine", _hmachines, key="hist_machine")

    with _hf2:
        _hkf_opts = ["All"] + sorted(
            df_qualified_hist["Key Failure"].dropna()
            .replace({"": "N/A", "nan": "N/A"}).unique().tolist()
        ) if "Key Failure" in df_qualified_hist.columns else ["All"]
        _hfkf = st.selectbox(" Key Failure", _hkf_opts, key="hist_kf")

    with _hf3:
        _hshift_opts = ["All"] + sorted(
            df_qualified_hist["Shift"].dropna()
            .replace({"": "N/A"}).unique().tolist()
        ) if "Shift" in df_qualified_hist.columns else ["All"]
        _hfsh = st.selectbox(" Shift", _hshift_opts, key="hist_shift")

    _dh = df_qualified_hist.copy()
    if _hfm != "All" and COL_MACHINE in _dh.columns:
        _dh = _dh[_dh[COL_MACHINE] == _hfm]
    if _hfkf != "All" and "Key Failure" in _dh.columns:
        _dh = _dh[_dh["Key Failure"].astype(str).str.strip() == _hfkf]
    if _hfsh != "All" and "Shift" in _dh.columns:
        _dh = _dh[_dh["Shift"].astype(str).str.strip() == _hfsh]

    st.markdown(
        f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:9px;'
        f'color:#9A7A60;margin-bottom:8px;letter-spacing:1px">'
        f'Showing <strong style="color:{TE_ORANGE}">{len(_dh)}</strong>'
        f' qualified stop(s)</div>',
        unsafe_allow_html=True)

    # ── QUALIFIED STOPS TABLE ──
    st.markdown(f'<div class="te-section"> QUALIFIED STOPS TABLE</div>', unsafe_allow_html=True)

    _hist_disp_cols = [c for c in [
        COL_MACHINE, COL_DATE, COL_STATUS, "mttr_h",
        "User ID", "Shift", "Key Failure",
        "Issue Description", "Action Taken",
        "Spare Part Ref", "Qty", "Unit Price (€)", "Total Part Cost",
    ] if c in _dh.columns]

    _dh_show = _dh[_hist_disp_cols].copy()
    if COL_DATE in _dh_show.columns:
        _dh_show[COL_DATE] = pd.to_datetime(
            _dh_show[COL_DATE], errors="coerce"
        ).dt.strftime("%m/%d/%Y").fillna("—")
    if "mttr_h" in _dh_show.columns:
        _dh_show["mttr_h"] = pd.to_numeric(_dh_show["mttr_h"], errors="coerce").round(4)

    st.dataframe(
        _dh_show.reset_index(drop=True),
        use_container_width=True, hide_index=True,
        height=min(700, max(300, len(_dh_show) * 40 + 52))
    )

    # ── Charts ──
    if len(_dh) >= 2:
        st.markdown(f'<div class="te-section">ANALYSIS</div>', unsafe_allow_html=True)
        _ch1, _ch2 = st.columns(2)

        with _ch1:
            st.markdown('<div class="chart-card"><div class="chart-header">'
                        '<div class="chart-dot"></div>'
                        '<div class="chart-title">Top Key Failures</div></div>',
                        unsafe_allow_html=True)
            if "Key Failure" in _dh.columns:
                _kf_ct = (
                    _dh["Key Failure"].replace({"": "N/A", "nan": "N/A"})
                    .value_counts().head(10).reset_index()
                )
                _kf_ct.columns = ["Key Failure", "Count"]
                _fig_kf = go.Figure(go.Bar(
                    x=_kf_ct["Count"], y=_kf_ct["Key Failure"], orientation="h",
                    marker=dict(color=[TE_ORANGE if i == 0 else TE_BLACK
                                       for i in range(len(_kf_ct))],
                                line=dict(width=0)),
                    hovertemplate="<b>%{y}</b><br>Count: %{x}<extra></extra>"
                ))
                apply(_fig_kf, height=320, showlegend=False,
                      margin=dict(l=10, r=10, t=20, b=20),
                      xaxis=dict(gridcolor="#F0E8E0", tickfont=dict(size=9, color="#9A7A60")),
                      yaxis=dict(showgrid=False, tickfont=dict(size=10, color="#4A3020")))
                st.plotly_chart(_fig_kf, use_container_width=True, config=PCONF)
            st.markdown("</div>", unsafe_allow_html=True)

        with _ch2:
            st.markdown('<div class="chart-card"><div class="chart-header">'
                        '<div class="chart-dot"></div>'
                        '<div class="chart-title">Stops per Machine</div></div>',
                        unsafe_allow_html=True)
            if COL_MACHINE in _dh.columns:
                _mac_ct = _dh[COL_MACHINE].value_counts().reset_index()
                _mac_ct.columns = ["Machine", "Stops"]
                _fig_mac = go.Figure(go.Bar(
                    x=_mac_ct["Machine"], y=_mac_ct["Stops"],
                    marker=dict(color=[PALETTE[i % len(PALETTE)]
                                       for i in range(len(_mac_ct))],
                                line=dict(width=0)),
                    hovertemplate="<b>%{x}</b><br>Stops: %{y}<extra></extra>"
                ))
                apply(_fig_mac, height=320, showlegend=False,
                      margin=dict(l=10, r=10, t=20, b=20),
                      xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#4A3020")),
                      yaxis=dict(gridcolor="#F0E8E0", tickfont=dict(size=9, color="#9A7A60")))
                st.plotly_chart(_fig_mac, use_container_width=True, config=PCONF)
            st.markdown("</div>", unsafe_allow_html=True)

    # ── EXPORT HISTORY ──
    st.markdown(f'<div class="te-section">EXPORT HISTORY</div>', unsafe_allow_html=True)
    _exp1, _exp2 = st.columns(2)
    _ts_hist = datetime.now().strftime("%Y%m%d_%H%M")
    with _exp1:
        st.download_button(
            "CSV  QUALIFIED HISTORY",
            data=_dh[_hist_disp_cols].to_csv(index=False, sep=";").encode("utf-8"),
            file_name=f"TE_History_Qualified_{_ts_hist}.csv",
            mime="text/csv", use_container_width=True
        )
    with _exp2:
        st.download_button(
            "CSV  FULL PERSISTENT DATA",
            data=df_hist.to_csv(index=False, sep=";").encode("utf-8"),
            file_name=f"TE_Persistent_Full_{_ts_hist}.csv",
            mime="text/csv", use_container_width=True
        )

    # ── RESET HISTORY ──
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="te-section">DANGER ZONE</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div style="background:#1a0a0a;border:2px solid {TE_RED};border-radius:12px;
                padding:18px 24px;margin-bottom:12px">
      <div style="font-family:'Barlow Condensed',sans-serif;font-size:16px;font-weight:800;
                  color:{TE_RED};text-transform:uppercase;letter-spacing:1px;margin-bottom:6px">
         Reset History
      </div>
      <div style="font-size:12px;color:#C0A080;line-height:1.7">
        This will permanently delete <strong style="color:{TE_WHITE}">{PERSISTENT_CSV}</strong>
        and all {len(df_hist):,} rows of saved data. This action <strong style="color:{TE_RED}">cannot be undone</strong>.
      </div>
    </div>
    """, unsafe_allow_html=True)

    if "confirm_reset" not in st.session_state:
        st.session_state.confirm_reset = False

    _r1, _r2, _r3 = st.columns([2, 1.5, 2])
    with _r2:
        if st.button(" Reset History", key="btn_reset_trigger",
                     help="Permanently delete all persistent data"):
            st.session_state.confirm_reset = True

    if st.session_state.confirm_reset:
        st.warning(
            f" **Are you sure?**  \n"
            f"This will delete **{len(df_hist):,} rows** from `{PERSISTENT_CSV}`.  \n"
            f"This action **cannot be undone**."
        )
        _confirm1, _confirm2, _confirm3 = st.columns([2, 1, 2])
        with _confirm1:
            if st.button(" Yes, Reset Everything", key="btn_reset_confirm",
                         help="Permanently reset all persistent data"):
                try:
                    # Windows-safe: overwrite with empty DataFrame (same columns)
                    # instead of os.remove() which causes WinError 32 on locked files
                    _empty_cols = []
                    if os.path.exists(PERSISTENT_CSV):
                        try:
                            _empty_cols = pd.read_csv(
                                PERSISTENT_CSV, nrows=0, low_memory=False
                            ).columns.tolist()
                        except Exception:
                            pass
                    pd.DataFrame(columns=_empty_cols).to_csv(
                        PERSISTENT_CSV, index=False, encoding="utf-8"
                    )
                    st.session_state.confirm_reset = False
                    st.session_state.edited_df     = None
                    st.session_state.library_df    = None
                    st.success(" History reset. Persistent file cleared (columns preserved).")
                    st.rerun()
                except Exception as _e:
                    st.error(f" Could not reset: {_e}")
        with _confirm2:
            if st.button(" Cancel", key="btn_reset_cancel"):
                st.session_state.confirm_reset = False
                st.rerun()

    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: DASHBOARD (default)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  DATA SOURCE PRIORITY
#  1st: file in uploader  →  read, archive, save to persistent
#  2nd: library_df loaded via selectbox
#  3rd: tpm_data_persistent.csv (auto-load on refresh)
# ══════════════════════════════════════════════════════════════════════════════

# ── Resolve which source is active ───────────────────────────────────────────
_src = "none"
if uploaded is not None:
    _src = "upload"
elif st.session_state.get("library_df") is not None:
    _src = "library"
elif os.path.exists(PERSISTENT_CSV):
    _p = load_persistent()
    if not _p.empty:
        _src = "persistent"

# ── If nothing available → show welcome screen ───────────────────────────────
if _src == "none":
    st.markdown(f"""
    <div style="display:flex;justify-content:center;align-items:center;
                min-height:60vh;margin-top:20px">
    <div style="background:white;border:2px dashed {TE_ORANGE};border-radius:18px;
                padding:52px 50px;text-align:center;max-width:540px;
                box-shadow:0 6px 32px rgba(232,101,10,0.12)">
        <div style="display:inline-flex;align-items:center;gap:10px;
                    background:linear-gradient(135deg,{TE_ORANGE},{TE_DARK});
                    border-radius:10px;padding:10px 22px;
                    font-family:'Barlow Condensed',sans-serif;font-size:17px;
                    font-weight:900;letter-spacing:2.5px;color:white;
                    margin-bottom:24px;box-shadow:0 4px 16px rgba(232,101,10,0.40)">
            ≡ TE CONNECTIVITY
        </div>
        <div style="font-family:'Barlow Condensed',sans-serif;font-size:32px;
                    font-weight:800;color:{TE_BLACK};text-transform:uppercase;
                    letter-spacing:2px;margin-bottom:8px;line-height:1.1">
            STAMPING CMMS
        </div>
        <div style="width:50px;height:3px;
                    background:linear-gradient(90deg,{TE_ORANGE},{TE_ORANGE3});
                    border-radius:2px;margin:0 auto 18px auto"></div>
        <div style="font-size:13px;color:#9A7A60;margin-bottom:24px;line-height:1.8">
            Import your Hydra MES file or load from<br>
            <strong style="color:{TE_ORANGE}">📌 DATA LIBRARY</strong> in the sidebar.<br>
            <span style="font-size:11px;color:{TE_ORANGE};font-weight:600">
                ↑ Use the sidebar to import or restore data
            </span>
        </div>
        <div style="display:flex;gap:8px;justify-content:center;flex-wrap:wrap">
            <span style="background:#FFF0E6;border:1px solid {TE_ORANGE3};color:{TE_DARK};
                         font-size:11px;font-weight:700;border-radius:20px;
                         padding:5px 14px">.csv comma</span>
            <span style="background:#FFF0E6;border:1px solid {TE_ORANGE3};color:{TE_DARK};
                         font-size:11px;font-weight:700;border-radius:20px;
                         padding:5px 14px">.csv semicolon</span>
            <span style="background:#FFF0E6;border:1px solid {TE_ORANGE3};color:{TE_DARK};
                         font-size:11px;font-weight:700;border-radius:20px;
                         padding:5px 14px">.xlsx</span>
        </div>
    </div></div>
    """, unsafe_allow_html=True)
    st.stop()

# ──────────────────────────────────────────────────────────────────────────────
#  DATA PREP — load raw data from active source
# ──────────────────────────────────────────────────────────────────────────────

if _src == "upload":
    # ── Priority 1: Uploader ─────────────────────────────────────
    _is_new_file = (st.session_state.get("last_file") != uploaded.name)
    if _is_new_file:
        _df_new = load_data(uploaded)
        if not _df_new.empty and len(_df_new.columns) >= 2:
            # ✅ Parse succeeded
            _arc_name = archive_import(_df_new)
            if _arc_name:
                st.session_state["_archive_saved"] = _arc_name
            save_persistent(_df_new)
            st.session_state.library_df     = None
            st.session_state.library_active = None
            st.session_state.last_file      = uploaded.name
            st.session_state.edited_df      = None
            df_raw = _df_new.copy()
        else:
            # ⚠️ Parse failed — silently fall back to persistent, show soft warning
            _fb = load_persistent()
            if not _fb.empty:
                st.warning(
                    f"⚠️ Could not read **{uploaded.name}** (format issue). "
                    f"Showing last saved session instead.",
                    icon="⚠️"
                )
                df_raw = _fb
                _src = "persistent"
            else:
                st.error(
                    f"❌ Cannot read **{uploaded.name}**. "
                    "Please check the file format (CSV comma/semicolon or XLSX)."
                )
                st.stop()
    else:
        df_raw = load_data(uploaded)    # uses cached bytes — instant

elif _src == "library":
    # ── Priority 2: Library selectbox ────────────────────────────
    df_raw = st.session_state.library_df.copy()
    _lib_key = f"__lib__{st.session_state.get('library_active','')}"
    if st.session_state.get("last_file") != _lib_key:
        st.session_state.last_file = _lib_key
        st.session_state.edited_df = None

else:
    # ── Priority 3: Persistent CSV (auto-load on page refresh) ────
    df_raw = load_persistent()
    _pers_key = "__persistent__"
    if st.session_state.get("last_file") != _pers_key:
        st.session_state.last_file = _pers_key
        st.session_state.edited_df = None

df_raw = df_raw.loc[:, ~df_raw.columns.duplicated()]

missing = check_missing(df_raw)
if missing:
    st.error(
        f"**Missing columns :** `{'`, `'.join(missing)}`\n\n"
        f"**Columns found in file :**\n```\n{chr(10).join(df_raw.columns.tolist())}\n```"
    )
    st.stop()

for _qc, _qd in _QUAL_COLS_DEFAULTS:
    if _qc not in df_raw.columns:
        df_raw[_qc] = _qd

if st.session_state.edited_df is None:
    _df_persist = load_persistent()
    if not _df_persist.empty:
        df_raw = merge_qualifications(df_raw, _df_persist)
    st.session_state.edited_df = df_raw.copy()

_edf = st.session_state.edited_df
if "User ID" in _edf.columns:
    _blank_uid = _edf["User ID"].astype(str).str.strip().isin(["", "nan", "None"])
    _edf.loc[_blank_uid, "User ID"] = "TE"
for _qc, _qd in _QUAL_COLS_DEFAULTS:
    if _qc not in _edf.columns:
        _edf[_qc] = _qd
if "Qty" in _edf.columns and "Unit Price (€)" in _edf.columns:
    _edf["Total Part Cost"] = (
        pd.to_numeric(_edf["Qty"], errors="coerce").fillna(0) *
        pd.to_numeric(_edf["Unit Price (€)"], errors="coerce").fillna(0.0)
    ).round(2)
st.session_state.edited_df = _edf

df_raw = st.session_state.edited_df.copy()

df_raw[COL_MTTR] = pd.to_numeric(df_raw[COL_MTTR], errors="coerce").fillna(0.0)

has_mtbf = COL_MTBF in df_raw.columns
if has_mtbf:
    df_raw[COL_MTBF] = pd.to_numeric(df_raw[COL_MTBF], errors="coerce").fillna(0.0)
    if (df_raw[COL_MTBF] == 0).all():
        has_mtbf = False
else:
    df_raw[COL_MTBF] = 0.0

# Convert seconds to hours — these are raw per-event values
df_raw["mttr_h"] = df_raw[COL_MTTR] / 3600.0
df_raw["mtbf_h"] = df_raw[COL_MTBF] / 3600.0 if has_mtbf else 0.0

if "Manual Duration (min)" in df_raw.columns:
    dur_mask = df_raw["Manual Duration (min)"].notna() & (df_raw["Manual Duration (min)"] > 0)
    df_raw.loc[dur_mask, "mttr_h"] = df_raw.loc[dur_mask, "Manual Duration (min)"] / 60.0

# Date parsing
if COL_DATE in df_raw.columns:
    raw_dates = df_raw[COL_DATE].astype(str)
    parsed    = pd.Series([pd.NaT] * len(df_raw), dtype="datetime64[ns]")
    formats_to_try = [
        "%m/%d/%Y %H:%M", "%m/%d/%Y", "%m-%d-%Y %H:%M", "%m-%d-%Y",
        "%-m/%-d/%Y", "%#m/%#d/%Y",
    ]
    for fmt_str in formats_to_try:
        mask = parsed.isna()
        if not mask.any(): break
        try:
            parsed[mask] = pd.to_datetime(
                raw_dates[mask], format=fmt_str, errors="coerce", dayfirst=False)
        except Exception:
            pass
    mask = parsed.isna()
    if mask.any():
        parsed[mask] = pd.to_datetime(raw_dates[mask], errors="coerce", dayfirst=False)
    df_raw[COL_DATE] = parsed
    df_raw["date_only"] = df_raw[COL_DATE].dt.normalize()
    if df_raw["date_only"].notna().sum() == 0:
        st.warning(f" Column `{COL_DATE}`: no valid date parsed.")


# ──────────────────────────────────────────────────────────────────────────────
#  SIDEBAR FILTERS
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    all_machines = sorted(df_raw[COL_MACHINE].dropna().unique().tolist())
    sel_machines = st.multiselect(
        "Machines", options=all_machines, default=all_machines, placeholder="Select…"
    )
    if "date_only" in df_raw.columns:
        valid_d = df_raw["date_only"].dropna()
        if len(valid_d):
            from datetime import date as dt_date
            dmin = valid_d.min().date()
            dmax = valid_d.max().date()
            dmax_cal = max(dmax, dt_date.today())
            dr = st.date_input("Period", value=(dmin, dmax),
                               min_value=dmin, max_value=dmax_cal,
                               format="DD/MM/YYYY")
        else:
            dr = None
    else:
        dr = None

    st.markdown("---")
    if _src == "upload" and uploaded is not None:
        _src_name = uploaded.name
    elif _src == "library":
        _src_name = os.path.basename(st.session_state.get("library_active", "Library"))
    else:
        _src_name = os.path.basename(PERSISTENT_CSV)
    st.markdown(f"""
    <div style="font-size:10px;color:rgba(255,255,255,0.3);
                font-family:'JetBrains Mono',monospace;letter-spacing:1px">
         {_src_name}<br>
         {len(df_raw):,} rows<br>
         Persistent: {"✓" if os.path.exists(PERSISTENT_CSV) else "none"}<br>
         Archive: {len(list_archive())} file(s)<br><br>
        TE CONNECTIVITY © {datetime.now().year}
    </div>
    """, unsafe_allow_html=True)

if not sel_machines:
    st.warning(" Please select at least one machine.")
    st.stop()

df = df_raw[df_raw[COL_MACHINE].isin(sel_machines)].copy()
if "date_only" in df.columns and dr and isinstance(dr, (list, tuple)) and len(dr) == 2:
    df = df[(df["date_only"].dt.date >= dr[0]) & (df["date_only"].dt.date <= dr[1])]

if df.empty:
    st.warning("No data for this selection.")
    st.stop()


# ──────────────────────────────────────────────────────────────────────────────
#  KPI CALCULATIONS — MEAN-BASED (MTTR = Mean Time To Repair, MTBF = Mean Time Between Failures)
#
#  For each machine:
#    nb_failures    = number of events where mttr_h > 0
#    mean_mttr      = Total Downtime / nb_failures           (MTTR definition)
#    mean_mtbf      = Total Uptime   / nb_failures           (MTBF definition)
#    availability   = mean_mtbf / (mean_mtbf + mean_mttr)   (standard formula)
#
#  Global KPIs aggregate means across all selected machines.
# ──────────────────────────────────────────────────────────────────────────────

stop_rows = df[df["mttr_h"] > 0].copy()
nb_arrets = len(stop_rows)  # number of failures

# Total downtime / uptime across all selected data
mt_total = df["mttr_h"].sum()          # total downtime hours
mb_total = df["mtbf_h"].sum()          # total uptime hours

# ── MEAN MTTR = total downtime / number of failures
mttr_mean_h = round(mt_total / nb_arrets, 4) if nb_arrets > 0 else 0.0

# ── MEAN MTBF = total uptime / number of failures (only if MTBF data exists)
if has_mtbf and nb_arrets > 0:
    mtbf_mean_h = round(mb_total / nb_arrets, 4)
else:
    mtbf_mean_h = 0.0

# ── AVAILABILITY = MTBF / (MTBF + MTTR)
if has_mtbf and (mtbf_mean_h + mttr_mean_h) > 0:
    dispo = round(mtbf_mean_h / (mtbf_mean_h + mttr_mean_h) * 100, 2)
    dispo_mode = "MTBF-based"
else:
    # Fallback: production status ratio
    prod_mask  = df[COL_STATUS].str.upper().str.contains("PRODUCTION", na=False)
    n_prod     = prod_mask.sum()
    n_total    = len(df)
    dispo      = round(n_prod / n_total * 100, 2) if n_total > 0 else 100.0
    dispo_mode = "Status-based"

# ── Per-machine aggregation with MEAN-based KPIs ──
def _machine_kpis(group):
    _failures = (group["mttr_h"] > 0).sum()
    _total_dt = group["mttr_h"].sum()           # total downtime
    _total_ut = group["mtbf_h"].sum()           # total uptime
    _mean_mttr = _total_dt / _failures if _failures > 0 else 0.0
    _mean_mtbf = _total_ut / _failures if (_failures > 0 and has_mtbf) else 0.0
    if has_mtbf and (_mean_mtbf + _mean_mttr) > 0:
        _avail = round(_mean_mtbf / (_mean_mtbf + _mean_mttr) * 100, 2)
    else:
        _prod_n = group[COL_STATUS].str.upper().str.contains("PRODUCTION", na=False).sum()
        _avail  = round(_prod_n / len(group) * 100, 2) if len(group) > 0 else 100.0
    return pd.Series({
        "mean_mttr_h": round(_mean_mttr, 4),
        "mean_mtbf_h": round(_mean_mtbf, 4),
        "nb_failures": int(_failures),
        "nb_events":   len(group),
        "dispo":       _avail,
    })

ma = df.groupby(COL_MACHINE).apply(_machine_kpis).reset_index()

# Pareto: total downtime per machine (for ranking worst actors)
pareto = df[df["mttr_h"] > 0].groupby(COL_MACHINE)["mttr_h"].sum().reset_index()
pareto.columns = ["Machine", "MTTR_total_h"]
pareto = pareto.sort_values("MTTR_total_h", ascending=False).reset_index(drop=True)
pareto["Pct"]   = (pareto["MTTR_total_h"] / pareto["MTTR_total_h"].sum() * 100).round(1)
pareto["Cumul"] = pareto["Pct"].cumsum().round(1)

kpi = dict(
    dispo=dispo, mttr_mean_h=mttr_mean_h, mtbf_mean_h=mtbf_mean_h,
    nb_arrets=nb_arrets, mttr_total_h=round(mt_total, 2),
    mtbf_total_h=round(mb_total, 2), nb_rows=len(df),
    by_machine=ma, pareto=pareto,
)

has_prod = COL_PROD in df.columns
if has_prod:
    prod_ct    = df[df[COL_PROD].str.lower().str.contains("prod", na=False)].shape[0]
    nonprod_ct = len(df) - prod_ct
else:
    prod_ct = nonprod_ct = 0


# ──────────────────────────────────────────────────────────────────────────────
#  HEADER
# ──────────────────────────────────────────────────────────────────────────────
dispo_col = TE_GREEN if dispo >= 95 else TE_AMBER if dispo >= 90 else TE_RED
dispo_lbl = "On Target " if dispo >= 95 else "Watch " if dispo >= 90 else "Critical "

st.markdown(f"""
<div class="te-header">
  <div style="display:flex;justify-content:space-between;align-items:flex-start">
    <div>
      <div class="te-header-tag">Stamping Department · Bruderer Presses</div>
      <div class="te-header-title">Stamping <span>CMMS</span></div>
      <div class="te-header-sub">Mean MTTR · Mean MTBF · Availability · Criticality · Pareto Analysis</div>
      <div class="te-header-badge">
         {df[COL_MACHINE].nunique()} machine{"s" if df[COL_MACHINE].nunique()>1 else ""}
        &nbsp;·&nbsp; {len(df):,} events
        &nbsp;·&nbsp; <span style="color:{dispo_col};font-weight:700">{dispo}% — {dispo_lbl}</span>
      </div>
    </div>
    <div class="te-header-right">
      <div class="te-live"><div class="te-live-dot"></div>SYSTEM ACTIVE</div>
      <div style="font-family:'JetBrains Mono',monospace;font-size:10px;
                  color:rgba(255,255,255,0.35);text-align:right;margin-top:4px">
        TANGIER · PLANT 1310
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if not has_mtbf:
    st.markdown(f"""
    <div style="background:#fff8e1;border:1px solid #ffe082;border-left:4px solid {TE_AMBER};
                border-radius:8px;padding:10px 18px;font-size:12px;color:#5d4037;
                margin-bottom:12px;display:flex;align-items:center;gap:10px">
         <span><strong>MTBF column absent or empty.</strong>
      Availability computed from machine status (PRODUCTION rows / total).</span>
    </div>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
#  TABS
# ──────────────────────────────────────────────────────────────────────────────
tab_kpi, tab_qual = st.tabs([
    "  Performance Analysis (KPIs)",
    "  Stops Qualification",
])


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 — KPI ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with tab_kpi:

    st.markdown(f"""
    <div class="te-statusbar">
      <div class="te-dot-green"></div>
      <div class="te-statusbar-item"><strong>{len(df):,}</strong>&nbsp;events</div>
      <div class="te-sep"></div>
      <div class="te-statusbar-item">Failures: <strong>{nb_arrets}</strong></div>
      <div class="te-sep"></div>
      <div class="te-statusbar-item">Machines: <strong>{df[COL_MACHINE].nunique()}</strong></div>
      {"<div class='te-sep'></div><div class='te-statusbar-item'>Production: <strong>"+str(prod_ct)+"</strong></div>" if has_prod else ""}
      <div class="te-sep"></div>
      <div class="te-statusbar-item">
        Availability: <strong style="color:{dispo_col}">{dispo}% — {dispo_lbl}</strong>
      </div>
      <div class="te-sep"></div>
      <div class="te-statusbar-item" style="font-size:10px;color:#9A7A60">
        Formula: MTBF/(MTBF+MTTR)
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── KPI Cards ──
    _total_cost = 0.0
    if st.session_state.edited_df is not None and "Total Part Cost" in st.session_state.edited_df.columns:
        _total_cost = float(
            pd.to_numeric(st.session_state.edited_df["Total Part Cost"], errors="coerce")
            .fillna(0).sum())

    st.markdown('<div class="te-section">⊞ Main KPIs</div>', unsafe_allow_html=True)
    c1, c2, c3, c4, c5 = st.columns(5)

    for col, label, value, unit, is_cost in [
        (c1, "AVAILABILITY",
         f"{dispo}%",
         f"MTBF/(MTBF+MTTR)" if has_mtbf else "Status-based",
         False),
        (c2, "MEAN MTTR",
         f"{mttr_mean_h:.3f} h",
         f"{round(mttr_mean_h*60,1)} min · Total DT / Failures",
         False),
        (c3, "MEAN MTBF",
         f"{mtbf_mean_h:.2f} h" if has_mtbf else "N/A",
         "Total UT / Failures" if has_mtbf else "Column absent",
         False),
        (c4, "TOTAL FAILURES",
         f"{nb_arrets}",
         f"Out of {len(df):,} events",
         False),
        (c5, "MAINT. COST",
         f"€ {_total_cost:,.2f}",
         "Spare parts (qualified stops)",
         True),
    ]:
        with col:
            _cstyle = ("background:linear-gradient(135deg,#1a3a1a,#0d200d);"
                       "border:1.5px solid #27AE60"
                       if is_cost and _total_cost > 0 else "")
            st.markdown(f"""
            <div class="kpi-card" style="{_cstyle}">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{value}</div>
              <div class="kpi-divider"></div>
              <div class="kpi-unit">{unit}</div>
            </div>""", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────────────────────
    #  PERFORMANCE TREND — Mean-based aggregation per period
    # ──────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="te-section"> Performance Trend</div>', unsafe_allow_html=True)

    if "date_only" in df.columns:
        _MONTH_FR = {1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"May", 6:"Jun",
                     7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec"}
        _df_t = df.copy()
        _dt   = pd.to_datetime(_df_t["date_only"], errors="coerce")

        _df_t["_month_year_n"] = _dt.dt.year.astype("Int64") * 100 + _dt.dt.month.astype("Int64")
        _df_t["_month_lbl"]    = (_dt.dt.month.map(_MONTH_FR).fillna("—") +
                                   " " + _dt.dt.year.astype(str).str[-2:])
        _df_t["_week_year_n"]  = (_dt.dt.isocalendar().year.astype("Int64") * 100 +
                                   _dt.dt.isocalendar().week.astype("Int64"))
        _df_t["_week_lbl"]     = "W" + _dt.dt.isocalendar().week.astype(str)

        def _te_agg_mean(lbl_col, sort_key):
            """
            Aggregate per period using MEAN-based MTTR/MTBF:
              mean_mttr = total_downtime / nb_failures
              mean_mtbf = total_uptime   / nb_failures
              availability = mean_mtbf / (mean_mtbf + mean_mttr)
            """
            grp = _df_t.groupby(lbl_col, as_index=False).agg(
                _sort_key=(sort_key, "first"),
                total_dt=("mttr_h", "sum"),
                total_ut=("mtbf_h", "sum"),
                nb_failures=("mttr_h", lambda x: (x > 0).sum()),
                nb_events=("mttr_h", "count"),
            ).sort_values("_sort_key")

            grp["mean_mttr_h"] = (grp["total_dt"] / grp["nb_failures"].replace(0, np.nan)).fillna(0).round(4)
            grp["mean_mtbf_h"] = (grp["total_ut"] / grp["nb_failures"].replace(0, np.nan)).fillna(0).round(4) if has_mtbf else 0.0

            if has_mtbf:
                grp["dispo"] = grp.apply(
                    lambda r: round(r.mean_mtbf_h / (r.mean_mtbf_h + r.mean_mttr_h) * 100, 2)
                    if (r.mean_mtbf_h + r.mean_mttr_h) > 0 else 100.0, axis=1)
            else:
                _prod = (_df_t[_df_t[COL_STATUS].str.upper()
                               .str.contains("PRODUCTION", na=False)]
                         .groupby(lbl_col).size().reset_index(name="n_prod"))
                _tot  = _df_t.groupby(lbl_col).size().reset_index(name="n_tot")
                _rat  = _prod.merge(_tot, on=lbl_col, how="right").fillna(0)
                _rat["dispo"] = (_rat["n_prod"] / _rat["n_tot"] * 100).round(2)
                grp = grp.merge(_rat[[lbl_col, "dispo"]], on=lbl_col, how="left")
                grp["dispo"] = grp["dispo"].fillna(0.0)

            return grp.rename(columns={lbl_col: "label"})[
                ["label", "mean_mttr_h", "mean_mtbf_h", "dispo",
                 "nb_failures", "nb_events"]
            ].reset_index(drop=True)

        _df_week  = _te_agg_mean("_week_lbl",  "_week_year_n")
        _df_month = _te_agg_mean("_month_lbl", "_month_year_n")

        def _te_line(x_vals, y_vals, title, y_title, color,
                     target=None, y_fmt=None, height=450):
            fig = go.Figure()
            if target is not None:
                fig.add_trace(go.Scatter(
                    x=x_vals, y=[target]*len(x_vals), mode="lines",
                    name=f"Target {target}%",
                    line=dict(color=TE_RED, dash="dot", width=1.8),
                    hoverinfo="skip"))
            fill_color = _hex_to_rgba(color, 0.08)
            fig.add_trace(go.Scatter(
                x=x_vals, y=y_vals, mode="lines+markers", name=y_title,
                line=dict(color=color, width=2.5),
                marker=dict(size=9, color=color,
                            line=dict(color="white", width=2), symbol="circle"),
                fill="tozeroy", fillcolor=fill_color,
                hovertemplate=f"<b>%{{x}}</b><br>{y_title}: <b>%{{y}}</b><extra></extra>",
            ))
            y_axis = dict(gridcolor="#F0E8E0", zeroline=False,
                          tickfont=dict(size=10, color="#9A7A60"))
            if y_fmt:
                y_axis["tickformat"] = y_fmt
            if target is not None:
                _safe_min = min((v for v in y_vals if pd.notna(v)), default=0)
                y_axis["range"] = [max(0, float(_safe_min) - 5), 105]
            apply(fig, height=height, showlegend=False,
                  title=dict(text=title,
                             font=dict(size=13, color=TE_BLACK, family="Barlow Condensed"),
                             x=0.01, y=0.97),
                  xaxis=dict(tickfont=dict(size=10, color="#9A7A60"),
                             gridcolor="#F0E8E0", zeroline=False,
                             tickangle=-40 if len(x_vals) > 10 else 0),
                  yaxis=y_axis,
                  margin=dict(l=12, r=12, t=40, b=40))
            return fig

        def _te_mini_label(txt):
            st.markdown(
                f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:8px;'
                f'font-weight:700;letter-spacing:2px;text-transform:uppercase;'
                f'color:{TE_ORANGE};margin:16px 0 6px 0;'
                f'display:flex;align-items:center;gap:8px">'
                f'<span style="width:12px;height:2px;background:{TE_ORANGE};'
                f'display:inline-block"></span>{txt}'
                f'<span style="flex:1;height:1px;'
                f'background:linear-gradient(90deg,#F0D0B0,transparent);'
                f'display:inline-block"></span></div>',
                unsafe_allow_html=True)

        def _te_recap_table(df_agg, periode_col):
            _tbl = df_agg.rename(columns={
                "label": "label_raw",
                "mean_mttr_h": "Mean MTTR (h)",
                "mean_mtbf_h": "Mean MTBF (h)",
                "dispo": "Availability (%)",
                "nb_failures": "Failures",
                "nb_events": "Events"
            })
            _tbl.insert(0, periode_col, _tbl.pop("label_raw"))
            def _sd(val):
                try:
                    v = float(val)
                    if v >= 95: return "background:#d5f5e3;color:#1e8449;font-weight:700"
                    if v >= 90: return "background:#fef9e7;color:#d68910;font-weight:700"
                    return              "background:#fdf2f2;color:#c0392b;font-weight:700"
                except: return ""
            st.dataframe(
                _tbl.style
                    .applymap(_sd, subset=["Availability (%)"])
                    .format({"Availability (%)": "{:.2f}%",
                             "Mean MTTR (h)": "{:.4f}",
                             "Mean MTBF (h)": "{:.4f}"}),
                use_container_width=True, hide_index=True,
                height=min(420, len(_tbl) * 36 + 42))

        _chart_choice = st.radio(
            " Select chart to display:",
            options=["Availability (%)", "Mean MTTR (h)", "Mean MTBF (h)"],
            index=0, horizontal=True, key="te_chart_pick")

        def _te_render_charts(df_v, is_mtbf, chart_key="Availability (%)"):
            x = df_v["label"].tolist()
            _specs = {
                "Availability (%)": (
                    df_v["dispo"].tolist(),
                    "Availability (%) — MTBF/(MTBF+MTTR)", "Avail. (%)", TE_GREEN, 95, ".1f"),
                "Mean MTBF (h)": (
                    df_v["mean_mtbf_h"].tolist() if is_mtbf else [0]*len(df_v),
                    "Mean MTBF (h) — Total Uptime / Failures", "Mean MTBF (h)", TE_NAVY, None, None),
                "Mean MTTR (h)": (
                    df_v["mean_mttr_h"].tolist(),
                    "Mean MTTR (h) — Total Downtime / Failures", "Mean MTTR (h)", TE_ORANGE, None, None),
            }
            _sel = chart_key if chart_key in _specs else "Availability (%)"
            _y, _ttl, _yt, _clr, _tgt, _fmt = _specs[_sel]
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.plotly_chart(
                _te_line(x, _y, _ttl, _yt, _clr, target=_tgt, y_fmt=_fmt, height=460),
                use_container_width=True, config=PCONF)
            if _sel == "Mean MTBF (h)" and not is_mtbf:
                st.caption(" MTBF column absent from Hydra file.")
            st.markdown("</div>", unsafe_allow_html=True)

        _stab_w, _stab_m = st.tabs(["  Weekly View", "  Monthly View"])

        with _stab_w:
            if len(_df_week) < 2:
                st.info("Not enough weekly data (minimum 2 weeks required).")
            else:
                _te_render_charts(_df_week, has_mtbf, _chart_choice)
                _te_mini_label("Weekly Summary")
                _te_recap_table(_df_week, "Week")

        with _stab_m:
            if len(_df_month) < 2:
                st.info("Not enough monthly data (minimum 2 months required).")
            else:
                _te_render_charts(_df_month, has_mtbf, _chart_choice)
                _te_mini_label("Monthly Summary")
                _te_recap_table(_df_month, "Month")

    else:
        st.info(" Column `plant_shift_date` absent — time trend unavailable.")

    # ── Pareto + Pie ──
    st.markdown('<div class="te-section"> Pareto & Cause Analysis</div>', unsafe_allow_html=True)
    col_l, col_r = st.columns(2, gap="medium")

    with col_l:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title"> Downtime Pareto</div></div>""",
          unsafe_allow_html=True)
        if not pareto.empty:
            bc = [TE_ORANGE if i < 2 else TE_NAVY if i < 4 else "#A8A8A8"
                  for i in range(len(pareto))]
            fig_par = make_subplots(specs=[[{"secondary_y": True}]])
            fig_par.add_trace(go.Bar(
                x=pareto["Machine"], y=pareto["MTTR_total_h"], name="Downtime (h)",
                marker=dict(color=bc, line=dict(width=0)),
                text=[f"{v:.2f}h" for v in pareto["MTTR_total_h"]],
                textposition="outside", textfont=dict(size=10, color="#4A3020"),
                hovertemplate="<b>%{x}</b><br>Total Downtime: %{y:.3f} h<extra></extra>"
            ), secondary_y=False)
            fig_par.add_trace(go.Scatter(
                x=pareto["Machine"], y=pareto["Cumul"], name="Cumul (%)",
                mode="lines+markers",
                line=dict(color=TE_RED, width=2.5),
                marker=dict(size=8, color=TE_RED, line=dict(color="white", width=2)),
                hovertemplate="<b>%{x}</b><br>Cumul: <b>%{y:.1f}%</b><extra></extra>"
            ), secondary_y=True)
            fig_par.add_hline(y=80, line_dash="dot", line_color=TE_RED, line_width=1.5,
                              secondary_y=True, annotation_text="80%",
                              annotation_font=dict(color=TE_RED, size=10))
            fig_par.update_layout(**{**PL, "height": 320, "bargap": 0.3,
                "yaxis":  dict(title="Total Downtime (h)", gridcolor="#F0E8E0",
                               tickfont=dict(size=9, color="#9A7A60"), zeroline=False),
                "yaxis2": dict(title="Cumul (%)", range=[0, 115], ticksuffix="%",
                               gridcolor="#F0E8E0", tickfont=dict(size=9, color="#9A7A60"),
                               zeroline=False),
                "xaxis":  dict(tickfont=dict(size=11, color="#4A3020"), zeroline=False),
            })
            st.plotly_chart(fig_par, use_container_width=True, config=PCONF)
            top1 = pareto.iloc[0]
            top2_pct = pareto.head(2)["Pct"].sum()
            st.markdown(
                f'<div class="te-insight-crit"> <strong>{top1["Machine"]}</strong> '
                f'= <strong>{top1["Pct"]}%</strong> of total downtime. '
                f'Top 2 machines = <strong>{top2_pct:.1f}%</strong>.</div>',
                unsafe_allow_html=True)
        else:
            st.info("No stops detected.")
        st.markdown("</div>", unsafe_allow_html=True)

    with col_r:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title"> Cause Analysis</div></div>""",
          unsafe_allow_html=True)
        sc = df[COL_STATUS].value_counts().reset_index()
        sc.columns = ["Statut", "Nombre"]
        status_colors = {"PRODUCTION": TE_GREEN, "micro stops": TE_ORANGE,
                         "Réparation Peripheri": TE_RED, "Réparation": TE_RED}
        pie_colors = [status_colors.get(s, "#A8A8A8") for s in sc["Statut"]]
        fig_pie = go.Figure(go.Pie(
            labels=sc["Statut"], values=sc["Nombre"], hole=0.58,
            marker=dict(colors=pie_colors, line=dict(color="white", width=3)),
            textinfo="percent", textfont=dict(size=11, family="Barlow Condensed"),
            hovertemplate="<b>%{label}</b><br>%{value} events<br>%{percent}<extra></extra>"
        ))
        apply(fig_pie, height=320,
            annotations=[dict(text=f"<b>{len(df):,}</b><br>events", x=0.5, y=0.5,
                showarrow=False, font=dict(size=14, color=TE_BLACK,
                                           family="Barlow Condensed"))],
            legend=dict(orientation="v", x=1, y=0.5, font=dict(size=10)))
        st.plotly_chart(fig_pie, use_container_width=True, config=PCONF)
        micro_row = sc[sc["Statut"].str.lower().str.contains("micro", na=False)]
        rep_row   = sc[sc["Statut"].str.lower().str.contains("réparat|reparat", na=False)]
        if not micro_row.empty and not rep_row.empty:
            mc = int(micro_row["Nombre"].sum())
            rc = int(rep_row["Nombre"].sum())
            dominant = "Micro-Stops" if mc > rc else "Repairs"
            st.markdown(
                f'<div class="te-insight"> <strong>{dominant}</strong> dominate '
                f'({max(mc,rc)} occurrences). '
                f'{"Focus on 5S and standardization." if mc > rc else "Strengthen preventive maintenance."}'
                f'</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # ── Criticality Matrix + Daily Trend ──
    st.markdown('<div class="te-section"> Criticality & Time Trend</div>', unsafe_allow_html=True)
    col_a, col_b = st.columns([2, 3], gap="medium")

    with col_a:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title"> Criticality Matrix</div></div>""",
          unsafe_allow_html=True)
        # Use mean MTTR / mean MTBF for criticality axes
        mx_v = ma["mean_mttr_h"].max() * 1.55 or 10
        my_v = ma["mean_mtbf_h"].max() * 1.55 or 100
        midx, midy = mx_v / 2, my_v / 2
        fig_mat = go.Figure()
        for x0, y0, x1, y1, c in [
            (0, midy, midx, my_v, "rgba(39,174,96,0.07)"),
            (midx, midy, mx_v, my_v, "rgba(27,42,74,0.07)"),
            (0, 0, midx, midy, "rgba(230,126,34,0.07)"),
            (midx, 0, mx_v, midy, "rgba(192,57,43,0.10)"),
        ]:
            fig_mat.add_shape(type="rect", x0=x0, y0=y0, x1=x1, y1=y1,
                              fillcolor=c, line_width=0, layer="below")
        for txt, x, y, tc in [
            ("↖ RELIABLE",  midx*0.04, my_v*0.97, TE_GREEN),
            ("↗ WATCH",     mx_v*0.54, my_v*0.97, TE_NAVY),
            ("↙ IMPROVE",   midx*0.04, my_v*0.47, TE_AMBER),
            ("↘ CRITICAL",  mx_v*0.54, my_v*0.47, TE_RED),
        ]:
            fig_mat.add_annotation(x=x, y=y, text=txt, showarrow=False,
                font=dict(size=9, color=tc, family="Barlow Condensed"),
                xanchor="left", yanchor="top")
        fig_mat.add_hline(y=midy, line_dash="dot", line_color="#D4CFC9", line_width=1.5)
        fig_mat.add_vline(x=midx, line_dash="dot", line_color="#D4CFC9", line_width=1.5)
        for i, row in ma.iterrows():
            c_dot = PALETTE[i % len(PALETTE)]
            fig_mat.add_trace(go.Scatter(
                x=[row["mean_mttr_h"]], y=[row["mean_mtbf_h"]],
                mode="markers+text", name=row[COL_MACHINE],
                text=[row[COL_MACHINE]], textposition="top center",
                textfont=dict(size=11, color=c_dot, family="Barlow Condensed"),
                marker=dict(size=min(60, max(22, int(row["nb_failures"]) * 3)),
                            color=c_dot, opacity=0.88,
                            line=dict(color="white", width=3)),
                hovertemplate=(f"<b>{row[COL_MACHINE]}</b><br>"
                               "Mean MTTR: %{x:.3f} h<br>Mean MTBF: %{y:.3f} h<br>"
                               f"Avail.: {row['dispo']}%<extra></extra>")
            ))
        apply(fig_mat, height=360, showlegend=False,
            xaxis=dict(title="Mean MTTR (h)", range=[0, mx_v], gridcolor="#F0E8E0",
                       tickfont=dict(size=9, color="#9A7A60"), zeroline=False),
            yaxis=dict(title="Mean MTBF (h)", range=[0, my_v], gridcolor="#F0E8E0",
                       tickfont=dict(size=9, color="#9A7A60"), zeroline=False))
        st.plotly_chart(fig_mat, use_container_width=True, config=PCONF)
        st.markdown("""
        <div class="quad-grid">
          <div class="quad q-good"><h5>↖ Reliable + Fast</h5><p>Maintain standard PM</p></div>
          <div class="quad q-watch"><h5>↗ Monitor</h5><p>Improve repair procedure</p></div>
          <div class="quad q-warn"><h5>↙ Improve</h5><p>Reinforced preventive PM</p></div>
          <div class="quad q-crit"><h5>↘ Bad Actor </h5><p>Absolute TPM priority</p></div>
        </div>""", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_b:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title"> Daily Availability Trend</div></div>""",
          unsafe_allow_html=True)
        if "date_only" in df.columns:
            # Daily availability using mean-based formula
            def _daily_avail(grp):
                _f = (grp["mttr_h"] > 0).sum()
                _dt = grp["mttr_h"].sum()
                _ut = grp["mtbf_h"].sum()
                _m_mttr = _dt / _f if _f > 0 else 0.0
                _m_mtbf = _ut / _f if (_f > 0 and has_mtbf) else 0.0
                if (_m_mtbf + _m_mttr) > 0:
                    return round(_m_mtbf / (_m_mtbf + _m_mttr) * 100, 1)
                return 100.0

            dm = df.groupby(["date_only", COL_MACHINE]).apply(_daily_avail).reset_index()
            dm.columns = ["date_only", COL_MACHINE, "dp"]
            da = df.groupby("date_only").apply(_daily_avail).reset_index()
            da.columns = ["date_only", "dp"]

            fig_evo = go.Figure()
            fig_evo.add_trace(go.Scatter(
                x=da["date_only"], y=[95]*len(da), mode="lines", name="Target 95%",
                line=dict(color=TE_RED, dash="dot", width=1.5)))
            for i, mac in enumerate(sorted(dm[COL_MACHINE].unique())):
                d2 = dm[dm[COL_MACHINE] == mac].sort_values("date_only")
                c2 = PALETTE[i % len(PALETTE)]
                fig_evo.add_trace(go.Scatter(
                    x=d2["date_only"], y=d2["dp"], mode="lines+markers", name=mac,
                    line=dict(color=c2, width=2),
                    marker=dict(size=6, color=c2, line=dict(color="white", width=2)),
                    hovertemplate=f"<b>{mac}</b><br>%{{x|%m/%d/%Y}}<br>Avail.: <b>%{{y}}%</b><extra></extra>"
                ))
            fig_evo.add_trace(go.Scatter(
                x=da["date_only"], y=da["dp"], mode="lines", name="⊞ Global",
                line=dict(color=TE_NAVY, width=3, dash="dot"),
                hovertemplate="Global<br>%{x|%m/%d/%Y}<br>Avail.: <b>%{y}%</b><extra></extra>"))
            apply(fig_evo, height=360,
                yaxis=dict(ticksuffix="%", range=[60, 105], gridcolor="#F0E8E0",
                           tickfont=dict(size=9, color="#9A7A60"), zeroline=False),
                xaxis=dict(tickformat="%d/%m", gridcolor="#F0E8E0",
                           tickfont=dict(size=9, color="#9A7A60"), zeroline=False))
            st.plotly_chart(fig_evo, use_container_width=True, config=PCONF)
        else:
            st.info("Column `plant_shift_date` absent — trend unavailable.")
        st.markdown("</div>", unsafe_allow_html=True)

    # ── Mean MTBF + Mean MTTR by Machine ──
    st.markdown('<div class="te-section"> Mean MTTR & Mean MTBF by Machine</div>', unsafe_allow_html=True)
    col_c, col_d = st.columns(2, gap="medium")

    with col_c:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title">Mean MTBF per Machine (h)</div></div>""",
          unsafe_allow_html=True)
        mb_m = ma[[COL_MACHINE, "mean_mtbf_h"]].sort_values("mean_mtbf_h", ascending=True)
        fig_mb = go.Figure(go.Bar(
            x=mb_m["mean_mtbf_h"], y=mb_m[COL_MACHINE], orientation="h",
            marker=dict(color=mb_m["mean_mtbf_h"],
                        colorscale=[[0,"#FAD9B5"],[0.5,TE_ORANGE2],[1,TE_DARK]],
                        showscale=False, line=dict(width=0)),
            text=mb_m["mean_mtbf_h"].apply(lambda v: f"{v:.3f}h"),
            textposition="outside", textfont=dict(size=10, color="#6A4030"),
            hovertemplate="<b>%{y}</b><br>Mean MTBF: %{x:.4f} h<extra></extra>"
        ))
        apply(fig_mb, height=max(240, len(mb_m)*55), bargap=0.35, showlegend=False,
            xaxis=dict(gridcolor="#F0E8E0", tickfont=dict(size=9, color="#9A7A60"),
                       zeroline=False, title="Mean MTBF (h)"),
            yaxis=dict(showgrid=False, tickfont=dict(size=11, color="#4A3020")))
        st.plotly_chart(fig_mb, use_container_width=True, config=PCONF)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_d:
        st.markdown("""<div class="chart-card">
          <div class="chart-header"><div class="chart-dot"></div>
          <div class="chart-title">Mean MTTR per Machine (h)</div></div>""",
          unsafe_allow_html=True)
        mt_m = ma[[COL_MACHINE, "mean_mttr_h"]].sort_values("mean_mttr_h", ascending=False)
        fig_mt = go.Figure(go.Bar(
            x=mt_m[COL_MACHINE], y=mt_m["mean_mttr_h"],
            marker=dict(color=mt_m["mean_mttr_h"],
                        colorscale=[[0,"#FAD9B5"],[0.5,TE_ORANGE],[1,TE_RED]],
                        showscale=False, line=dict(width=0)),
            hovertemplate="<b>%{x}</b><br>Mean MTTR: %{y:.4f} h<extra></extra>"
        ))
        apply(fig_mt, height=max(240, len(mt_m)*55), bargap=0.35, showlegend=False,
            xaxis=dict(showgrid=False, tickfont=dict(size=11, color="#4A3020")),
            yaxis=dict(gridcolor="#F0E8E0", tickfont=dict(size=9, color="#9A7A60"),
                       zeroline=False, title="Mean MTTR (h)"))
        st.plotly_chart(fig_mt, use_container_width=True, config=PCONF)
        st.markdown("</div>", unsafe_allow_html=True)

    # ── Summary Table ──
    st.markdown('<div class="te-section"> Summary Table by Machine</div>', unsafe_allow_html=True)

    ma_disp = ma.rename(columns={
        COL_MACHINE:      "Machine",
        "mean_mttr_h":    "Mean MTTR (h)",
        "mean_mtbf_h":    "Mean MTBF (h)",
        "nb_failures":    "Failures",
        "nb_events":      "Events",
        "dispo":          "Availability (%)",
    })[["Machine","Mean MTTR (h)","Mean MTBF (h)","Failures","Events","Availability (%)"]].copy()

    worst = ma_disp.loc[ma_disp["Availability (%)"].idxmin()]
    best  = ma_disp.loc[ma_disp["Availability (%)"].idxmax()]
    ci1, ci2 = st.columns(2)
    with ci1:
        if worst["Availability (%)"] < 90:
            st.markdown(
                f'<div class="te-insight-crit"> <strong>Bad Actor: {worst["Machine"]}</strong>'
                f' — Avail. {worst["Availability (%)"]:.1f}% '
                f'(Mean MTTR = {worst["Mean MTTR (h)"]:.3f} h). Priority TPM action.</div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div class="te-insight-ok"> All machines ≥ 90%. '
                f'Best: <strong>{best["Machine"]}</strong> ({best["Availability (%)"]:.1f}%).</div>',
                unsafe_allow_html=True)
    with ci2:
        st.markdown(
            f'<div class="te-insight"> <strong>Mean MTTR (global):</strong> '
            f'{mttr_mean_h:.3f} h ({round(mttr_mean_h*60,1)} min) · '
            f'<strong>Total Failures:</strong> {nb_arrets} / {len(df):,} events · '
            f'Formula: Mean MTTR = Total Downtime / Failures</div>',
            unsafe_allow_html=True)

    def style_dispo(val):
        try:
            v = float(str(val).replace("%",""))
            if v >= 95: return "background-color:#d5f5e3;color:#1e8449;font-weight:700"
            if v >= 90: return "background-color:#fef9e7;color:#d68910;font-weight:700"
            return              "background-color:#fdf2f2;color:#c0392b;font-weight:700"
        except Exception: return ""

    def style_mttr(val):
        try:
            v  = float(val)
            mx2 = float(ma_disp["Mean MTTR (h)"].max()) or 1.0
            ratio = min(v / mx2, 1.0)
            g   = int(255 - ratio * 160)
            b2  = int(255 - ratio * 210)
            txt = "#7a2005" if ratio > 0.6 else "#4a3020"
            return f"background-color:rgb(255,{g},{b2});color:{txt};font-weight:{'700' if ratio>0.6 else '400'}"
        except Exception: return ""

    st.dataframe(
        ma_disp.style
            .applymap(style_dispo, subset=["Availability (%)"])
            .applymap(style_mttr,  subset=["Mean MTTR (h)"])
            .format({"Mean MTTR (h)":"{:.4f}","Mean MTBF (h)":"{:.4f}",
                     "Availability (%)":"{:.1f}%"}),
        use_container_width=True, hide_index=True
    )

    # ── Export ──
    st.markdown('<div class="te-section">⬇ Data Export</div>', unsafe_allow_html=True)
    today_str = datetime.now().strftime("%Y%m%d_%H%M")

    # ── PDF Builder ──
    def build_pdf(df_in: pd.DataFrame, kpi_d: dict,
                  ma_table: pd.DataFrame, pareto_df: pd.DataFrame) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib           import colors as rlc
        from reportlab.lib.units     import cm
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors    import HexColor
        from reportlab.platypus      import (SimpleDocTemplate, Paragraph,
                                              Spacer, Table, TableStyle,
                                              HRFlowable, PageBreak,
                                              Image as RLImage)

        C_NAVY = HexColor("#1B2A4A"); C_OR   = HexColor("#E8650A")
        C_DARK = HexColor("#C04D05"); C_CREAM= HexColor("#FFF8F2")
        C_BR   = HexColor("#A07858"); C_WH   = rlc.white
        C_GR   = HexColor("#27AE60"); C_RED2 = HexColor("#C0392B")
        C_AM   = HexColor("#F39C12"); C_BG   = HexColor("#F7F4F0")
        _PAL   = ["#E8650A","#1B2A4A","#27AE60","#C0392B","#8E44AD","#2980B9"]

        W, H   = A4
        MARGIN = 1.8 * cm
        IW     = W - 2 * MARGIN
        buf    = io.BytesIO()

        _sty = getSampleStyleSheet()
        def ps(name, **kw):
            return ParagraphStyle(name, parent=_sty["Normal"], **kw)

        S_BRAND = ps("brand", fontSize=9,  textColor=C_WH,
                      fontName="Helvetica-Bold", letterSpacing=4, leading=14)
        S_TITLE = ps("ctit",  fontSize=30, textColor=C_WH,
                      fontName="Helvetica-Bold", leading=34)
        S_SUB   = ps("csub",  fontSize=12, textColor=C_BG,
                      fontName="Helvetica", leading=16)
        S_DATE  = ps("cdat",  fontSize=10, textColor=C_OR,
                      fontName="Helvetica-Bold", leading=14)
        S_SEC   = ps("sec",   fontSize=11, textColor=C_OR,
                      fontName="Helvetica-Bold", leading=14,
                      spaceBefore=12, spaceAfter=5)
        S_SSEC  = ps("ssec",  fontSize=9,  textColor=C_NAVY,
                      fontName="Helvetica-Bold", leading=12,
                      spaceBefore=8, spaceAfter=4)
        S_BODY  = ps("body",  fontSize=9,  textColor=HexColor("#3A2A1A"), leading=13)
        S_CAP   = ps("cap",   fontSize=8,  textColor=C_BR,
                      leading=11, spaceAfter=3, leftIndent=4)

        story = []

        def hr(color=C_OR, thick=1.2):
            return HRFlowable(width="100%", thickness=thick,
                               color=color, spaceAfter=7, spaceBefore=2)

        def section(txt):
            story.append(Spacer(1, 0.1*cm))
            story.append(Paragraph(txt, S_SEC))
            story.append(hr())

        def _tbl(data, cws, hbg=C_NAVY, hfg=C_WH, fs=8, extras=None):
            cmds = [
                ("BACKGROUND",    (0,0), (-1,0),  hbg),
                ("TEXTCOLOR",     (0,0), (-1,0),  hfg),
                ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE",      (0,0), (-1,-1), fs),
                ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("BOX",           (0,0), (-1,-1), 0.5, HexColor("#DDD0C8")),
                ("INNERGRID",     (0,0), (-1,-1), 0.3, HexColor("#EDE0D4")),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_CREAM, C_WH]),
            ]
            if extras:
                cmds += extras
            t = Table(data, colWidths=cws, repeatRows=1)
            t.setStyle(TableStyle(cmds))
            return t

        def fig_to_png(fig, w_px=720, h_px=300):
            try:
                import kaleido  # noqa
                return fig.to_image(format="png", width=w_px, height=h_px, scale=2)
            except Exception:
                return None

        def insert_fig(fig, caption="", w_px=720, h_px=300, img_w=None):
            if img_w is None:
                img_w = IW
            png = fig_to_png(fig, w_px, h_px)
            if png:
                story.append(RLImage(io.BytesIO(png),
                                      width=img_w, height=img_w * h_px / w_px))
                if caption:
                    story.append(Paragraph(caption, S_CAP))
            else:
                story.append(Paragraph(
                    f"[Chart unavailable ({caption}) — pip install kaleido]", S_CAP))
            story.append(Spacer(1, 0.2*cm))

        def _pdf_te_agg(df_src):
            if "date_only" not in df_src.columns:
                return pd.DataFrame(), pd.DataFrame()
            _MFR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                    7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
            _t  = df_src.copy()
            _dt = pd.to_datetime(_t["date_only"], errors="coerce")
            _t["_month_year_n"] = (_dt.dt.year.astype("Int64")*100 +
                                    _dt.dt.month.astype("Int64"))
            _t["_month_lbl"]    = (_dt.dt.month.map(_MFR).fillna("—") + " " +
                                    _dt.dt.year.astype(str).str[-2:])
            _t["_week_year_n"]  = (_dt.dt.isocalendar().year.astype("Int64")*100 +
                                    _dt.dt.isocalendar().week.astype("Int64"))
            _t["_week_lbl"]     = "W" + _dt.dt.isocalendar().week.astype(str)

            def _agg(lbl_col, sort_key):
                grp = _t.groupby(lbl_col, as_index=False).agg(
                    _sk=(sort_key, "first"),
                    total_dt=("mttr_h", "sum"),
                    total_ut=("mtbf_h", "sum"),
                    nb_failures=("mttr_h", lambda x: (x > 0).sum()),
                    nb_events=("mttr_h", "count"),
                ).sort_values("_sk")
                grp["mean_mttr_h"] = (grp["total_dt"] / grp["nb_failures"].replace(0, np.nan)).fillna(0).round(4)
                grp["mean_mtbf_h"] = (grp["total_ut"] / grp["nb_failures"].replace(0, np.nan)).fillna(0).round(4) if has_mtbf else 0.0
                if has_mtbf:
                    grp["dispo"] = grp.apply(
                        lambda r: round(r.mean_mtbf_h/(r.mean_mtbf_h+r.mean_mttr_h)*100, 2)
                        if (r.mean_mtbf_h+r.mean_mttr_h) > 0 else 100.0, axis=1)
                else:
                    _prod = (_t[_t[COL_STATUS].str.upper()
                                .str.contains("PRODUCTION", na=False)]
                             .groupby(lbl_col).size().reset_index(name="n_prod"))
                    _tot  = _t.groupby(lbl_col).size().reset_index(name="n_tot")
                    _rat  = _prod.merge(_tot, on=lbl_col, how="right").fillna(0)
                    _rat["dispo"] = (_rat["n_prod"]/_rat["n_tot"]*100).round(2)
                    grp = grp.merge(_rat[[lbl_col,"dispo"]], on=lbl_col, how="left")
                    grp["dispo"] = grp["dispo"].fillna(0.0)
                return grp.rename(columns={lbl_col:"label"})[
                    ["label","mean_mttr_h","mean_mtbf_h","dispo","nb_failures","nb_events"]
                ].reset_index(drop=True)
            return _agg("_week_lbl","_week_year_n"), _agg("_month_lbl","_month_year_n")

        def _make_trend_fig(df_v, col, title, color_hex, target=None, h_px=280):
            x  = df_v["label"].tolist()
            y  = df_v[col].tolist()
            fig = go.Figure()
            if target is not None and x:
                fig.add_trace(go.Scatter(
                    x=x, y=[target]*len(x), mode="lines",
                    line=dict(color="#C0392B", dash="dot", width=2), hoverinfo="skip"))
            _rgba_fill = _hex_to_rgba(color_hex, 0.08)
            fig.add_trace(go.Scatter(
                x=x, y=y, mode="lines+markers",
                line=dict(color=color_hex, width=2.5),
                marker=dict(size=8, color=color_hex, line=dict(color="white", width=1.8)),
                fill="tozeroy", fillcolor=_rgba_fill, hoverinfo="skip"))
            _ya = dict(gridcolor="#F0E8E0", zeroline=False, tickfont=dict(size=9))
            if target is not None and y:
                _ya["range"] = [max(0, min(float(v) for v in y)-5), 105]
            fig.update_layout(
                height=h_px, width=720, showlegend=False,
                paper_bgcolor="white", plot_bgcolor="#FAFAFA",
                title=dict(text=title, font=dict(size=12, color="#1B2A4A"), x=0.01),
                xaxis=dict(tickfont=dict(size=9), gridcolor="#F0E8E0",
                           tickangle=-35 if len(x) > 10 else 0),
                yaxis=_ya,
                margin=dict(l=48, r=16, t=36, b=48))
            return fig

        def _recap_pdf_table(df_v, periode_lbl):
            hdr = [[periode_lbl, "Mean MTTR (h)", "Mean MTBF (h)",
                    "Availability (%)", "Failures", "Events"]]
            rows = []
            extras = []
            for ri, row in df_v.iterrows():
                d = float(row["dispo"])
                rows.append([
                    str(row["label"]),
                    f"{float(row['mean_mttr_h']):.4f}",
                    f"{float(row['mean_mtbf_h']):.4f}",
                    f"{d:.2f}%",
                    str(int(row["nb_failures"])),
                    str(int(row["nb_events"])),
                ])
                ri_tbl = len(rows)
                bg = (HexColor("#D5F5E3") if d >= 95
                      else HexColor("#FEF9E7") if d >= 90
                      else HexColor("#FDEBD0"))
                tc = (HexColor("#1E8449") if d >= 95
                      else HexColor("#9A7D0A") if d >= 90
                      else HexColor("#922B21"))
                extras += [
                    ("BACKGROUND", (3, ri_tbl), (3, ri_tbl), bg),
                    ("TEXTCOLOR",  (3, ri_tbl), (3, ri_tbl), tc),
                    ("FONTNAME",   (3, ri_tbl), (3, ri_tbl), "Helvetica-Bold"),
                ]
            cws = [IW*x for x in [0.14, 0.17, 0.17, 0.22, 0.13, 0.13]]
            return _tbl(hdr + rows, cws, fs=8, extras=extras)

        # Cover page
        def _cover_row(content, bg, pad_top=8, pad_bot=8, left=10):
            t = Table([[content]], colWidths=[IW])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), bg),
                ("TOPPADDING",    (0,0),(-1,-1), pad_top),
                ("BOTTOMPADDING", (0,0),(-1,-1), pad_bot),
                ("LEFTPADDING",   (0,0),(-1,-1), left),
                ("RIGHTPADDING",  (0,0),(-1,-1), 10),
            ]))
            return t

        story.append(_cover_row(
            Paragraph("≡  TE CONNECTIVITY", S_BRAND),
            C_NAVY, pad_top=22, pad_bot=4))
        story.append(_cover_row(
            Paragraph("STAMPING DEPT  ·  BRUDERER PRESSES  ·  TANGIER PLANT 1310",
                      ps("bar", fontSize=8, textColor=C_WH,
                         fontName="Helvetica-Bold", letterSpacing=2)),
            C_OR, pad_top=6, pad_bot=6))

        _title_tbl = Table([[
            Paragraph("STAMPING CMMS", S_TITLE),
            Paragraph("Bruderer Presses S-001 → S-006<br/>Mean MTTR · Mean MTBF · Availability", S_SUB),
        ]], colWidths=[IW*0.60, IW*0.40])
        _title_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), C_NAVY),
            ("TOPPADDING",    (0,0),(-1,-1), 22),
            ("BOTTOMPADDING", (0,0),(-1,-1), 10),
            ("LEFTPADDING",   (0,0),(-1,-1), 10),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]))
        story.append(_title_tbl)
        story.append(_cover_row(
            Paragraph(f"REPORT GENERATED:  {datetime.now().strftime('%d/%m/%Y   %H:%M')}",
                      S_DATE),
            C_NAVY, pad_top=2, pad_bot=18))
        story.append(Spacer(1, 0.5*cm))

        _kpi_data = [
            ["INDICATOR", "VALUE"],
            ["Global Availability (MTBF/(MTBF+MTTR))", f"{kpi_d['dispo']:.2f} %"],
            ["Mean MTTR (Total DT / Failures)",         f"{kpi_d['mttr_mean_h']:.4f} h  ({round(kpi_d['mttr_mean_h']*60,1)} min)"],
            ["Mean MTBF (Total UT / Failures)",         f"{kpi_d['mtbf_mean_h']:.4f} h"],
            ["Total Failures",                          str(kpi_d['nb_arrets'])],
            ["Total Downtime",                          f"{kpi_d['mttr_total_h']:.2f} h"],
            ["Total Uptime",                            f"{kpi_d['mtbf_total_h']:.2f} h"],
            ["Total Events Analyzed",                   f"{kpi_d['nb_rows']:,}"],
        ]
        _cover_extras = []
        for _ri in range(1, len(_kpi_data)):
            _bg = HexColor("#FAD0A8") if _ri % 2 == 1 else C_CREAM
            _cover_extras.append(("BACKGROUND", (0,_ri),(0,_ri), _bg))
        story.append(_tbl(_kpi_data, [IW*0.62, IW*0.38], fs=9, extras=_cover_extras))
        story.append(PageBreak())

        # Page 2: KPIs + Pareto + Pie
        section("⊞ MAIN KPIs")
        _kw = IW / 4
        _kpi_card = Table([
            ["Availability", "Failures", "Mean MTTR", "Mean MTBF"],
            [f"{kpi_d['dispo']:.2f} %", str(kpi_d['nb_arrets']),
             f"{kpi_d['mttr_mean_h']:.4f} h", f"{kpi_d['mtbf_mean_h']:.4f} h"],
        ], colWidths=[_kw]*4)
        _kpi_card.setStyle(TableStyle([
            ("FONTNAME",     (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTNAME",     (0,1),(-1,1),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),(-1,0),  8),
            ("FONTSIZE",     (0,1),(-1,1),  15),
            ("TEXTCOLOR",    (0,0),(-1,-1), C_WH),
            ("ALIGN",        (0,0),(-1,-1), "CENTER"),
            ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",   (0,0),(-1,0),  6),
            ("BOTTOMPADDING",(0,0),(-1,0),  4),
            ("TOPPADDING",   (0,1),(-1,1),  8),
            ("BOTTOMPADDING",(0,1),(-1,1),  12),
            ("BACKGROUND",   (0,0),(0,-1),  C_NAVY),
            ("BACKGROUND",   (1,0),(1,-1),  C_OR),
            ("BACKGROUND",   (2,0),(2,-1),  C_DARK),
            ("BACKGROUND",   (3,0),(3,-1),  HexColor("#2A4A6A")),
            ("INNERGRID",    (0,0),(-1,-1), 1.5, C_WH),
        ]))
        story.append(_kpi_card)
        story.append(Spacer(1, 0.4*cm))

        section(" DOWNTIME PARETO")
        if not pareto_df.empty:
            _bc = [TE_ORANGE if i<2 else TE_NAVY if i<4 else "#A8A8A8"
                   for i in range(len(pareto_df))]
            _fp = make_subplots(specs=[[{"secondary_y": True}]])
            _fp.add_trace(go.Bar(
                x=pareto_df["Machine"], y=pareto_df["MTTR_total_h"],
                marker=dict(color=_bc, line=dict(width=0)),
                text=[f"{v:.2f}h" for v in pareto_df["MTTR_total_h"]],
                textposition="outside", hoverinfo="skip"),
                secondary_y=False)
            _fp.add_trace(go.Scatter(
                x=pareto_df["Machine"], y=pareto_df["Cumul"],
                mode="lines+markers",
                line=dict(color="#C0392B", width=2.5),
                marker=dict(size=8, color="#C0392B"), hoverinfo="skip"),
                secondary_y=True)
            _fp.add_hline(y=80, line_dash="dot", line_color="#C0392B",
                           line_width=1.5, secondary_y=True)
            _fp.update_layout(
                height=290, width=720, showlegend=False,
                paper_bgcolor="white", plot_bgcolor="#FAFAFA",
                margin=dict(l=44, r=44, t=16, b=44),
                xaxis=dict(tickfont=dict(size=10), gridcolor="#F0E8E0"),
                yaxis=dict(title="Total Downtime (h)", gridcolor="#F0E8E0",
                           tickfont=dict(size=9)),
                yaxis2=dict(title="Cumul (%)", range=[0,115], ticksuffix="%",
                             tickfont=dict(size=9)))
            insert_fig(_fp, "Downtime Pareto", w_px=720, h_px=290)

        story.append(PageBreak())

        # Page 3: Trend Analysis
        story.append(Paragraph("TREND ANALYSIS — MEAN MTTR / MEAN MTBF / AVAILABILITY", S_SEC))
        story.append(hr(thick=2))
        story.append(Spacer(1, 0.1*cm))
        _df_w, _df_m = _pdf_te_agg(df_in)
        _has_week  = not _df_w.empty and len(_df_w) >= 2
        _has_month = not _df_m.empty and len(_df_m) >= 2

        for _view_lbl, _df_v, _has_v, _periode_col in [
            ("Weekly View", _df_w, _has_week, "Week"),
            ("Monthly View", _df_m, _has_month, "Month"),
        ]:
            if not _has_v:
                story.append(Paragraph(f"[{_view_lbl} — insufficient data.]", S_CAP))
                continue
            story.append(Paragraph(_view_lbl, S_SSEC))
            story.append(hr(color=C_NAVY, thick=0.8))
            insert_fig(
                _make_trend_fig(_df_v, "dispo",
                                f"Availability (%) — {_view_lbl}  [MTBF/(MTBF+MTTR)]",
                                "#27AE60", target=95, h_px=250),
                f"Availability {_view_lbl.lower()}", w_px=720, h_px=250)
            insert_fig(
                _make_trend_fig(_df_v, "mean_mttr_h",
                                f"Mean MTTR (h) — {_view_lbl}  [Total DT / Failures]",
                                "#E8650A", h_px=230),
                f"Mean MTTR {_view_lbl.lower()}", w_px=720, h_px=230)
            if has_mtbf:
                insert_fig(
                    _make_trend_fig(_df_v, "mean_mtbf_h",
                                    f"Mean MTBF (h) — {_view_lbl}  [Total UT / Failures]",
                                    "#1B2A4A", h_px=230),
                    f"Mean MTBF {_view_lbl.lower()}", w_px=720, h_px=230)
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f"Summary Table — {_view_lbl}", S_SSEC))
            story.append(_recap_pdf_table(_df_v, _periode_col))
            story.append(Spacer(1, 0.4*cm))
        story.append(PageBreak())

        # Page 4: Summary Table by Machine
        section(" SUMMARY TABLE BY MACHINE")
        _hd_ma = [["Machine","Mean MTTR (h)","Mean MTBF (h)","Failures","Events","Availability (%)"]]
        _rows_ma = []
        _ext_ma  = []
        for _ri, _row in ma_table.iterrows():
            _d = float(_row["Availability (%)"])
            _rows_ma.append([
                str(_row["Machine"]),
                f"{float(_row['Mean MTTR (h)']):.4f}",
                f"{float(_row['Mean MTBF (h)']):.4f}",
                str(int(_row["Failures"])),
                str(int(_row["Events"])),
                f"{_d:.1f}%"])
            _ri_t = len(_rows_ma)
            _ebg  = (HexColor("#D5F5E3") if _d>=95
                     else HexColor("#FDEBD0") if _d>=90
                     else HexColor("#FADBD8"))
            _etc  = (HexColor("#1E8449") if _d>=95
                     else HexColor("#784212") if _d>=90
                     else HexColor("#922B21"))
            _ext_ma += [
                ("BACKGROUND",(5,_ri_t),(5,_ri_t),_ebg),
                ("TEXTCOLOR", (5,_ri_t),(5,_ri_t),_etc),
                ("FONTNAME",  (5,_ri_t),(5,_ri_t),"Helvetica-Bold"),
            ]
        story.append(_tbl(_hd_ma+_rows_ma,
                           [IW*x for x in [0.15,0.18,0.18,0.14,0.12,0.23]],
                           fs=9, extras=_ext_ma))
        story.append(Spacer(1, 0.3*cm))

        if not pareto_df.empty:
            section(" DOWNTIME PARETO — DETAIL")
            _hd_p  = [["Machine","Total Downtime (h)","Part (%)","Cumul (%)"]]
            _rows_p = [[str(r["Machine"]),
                         f"{float(r['MTTR_total_h']):.3f}",
                         f"{float(r['Pct']):.1f}%",
                         f"{float(r['Cumul']):.1f}%"]
                        for _,r in pareto_df.iterrows()]
            story.append(_tbl(_hd_p+_rows_p,
                               [IW*x for x in [0.30,0.28,0.21,0.21]], fs=9))

        story.append(PageBreak())

        # Page 5: Spare Parts
        story.append(PageBreak())
        _hdr_cost = Table([[
            Paragraph("  SPARE PARTS &amp; MAINTENANCE COSTS",
                      ps("sph", fontSize=16, fontName="Helvetica-Bold",
                         textColor=C_WH, leading=20)),
            Paragraph("Financial breakdown by event",
                      ps("sps", fontSize=9, textColor=HexColor("#FAD0A8"),
                         fontName="Helvetica", leading=13, alignment=2)),
        ]], colWidths=[IW * 0.65, IW * 0.35])
        _hdr_cost.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), C_NAVY),
            ("TOPPADDING",    (0,0),(-1,-1), 14),
            ("BOTTOMPADDING", (0,0),(-1,-1), 14),
            ("LEFTPADDING",   (0,0),(-1,-1), 14),
            ("RIGHTPADDING",  (0,0),(-1,-1), 14),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]))
        story.append(_hdr_cost)
        story.append(hr(color=C_OR, thick=3))
        story.append(Spacer(1, 0.2*cm))

        if "Total Part Cost" in df_in.columns:
            _cd = df_in[pd.to_numeric(df_in["Total Part Cost"], errors="coerce")
                        .fillna(0) > 0].copy()
        else:
            _cd = pd.DataFrame()

        _total_spend = float(
            pd.to_numeric(df_in.get("Total Part Cost", pd.Series([0.0])),
                          errors="coerce").fillna(0).sum()
        ) if "Total Part Cost" in df_in.columns else 0.0
        _n_cost_events = len(_cd)
        _n_parts_filled = int(
            _cd["Spare Part Ref"].astype(str).str.strip()
            .replace({"":"","nan":"","None":""}).ne("").sum()
        ) if "Spare Part Ref" in _cd.columns else 0

        _kpi_cost_data = [
            ["TOTAL SPEND", "EVENTS WITH COST", "PARTS REFERENCED"],
            [f"€ {_total_spend:,.2f}", str(_n_cost_events), str(_n_parts_filled)],
        ]
        _kc = Table(_kpi_cost_data, colWidths=[IW/3]*3)
        _kc.setStyle(TableStyle([
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0),  8),
            ("FONTNAME",      (0,1),(-1,1),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,1),(-1,1),  16),
            ("TEXTCOLOR",     (0,0),(-1,-1), C_WH),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0),(-1,0),  8),
            ("BOTTOMPADDING", (0,0),(-1,0),  4),
            ("TOPPADDING",    (0,1),(-1,1),  8),
            ("BOTTOMPADDING", (0,1),(-1,1),  12),
            ("BACKGROUND",    (0,0),(0,-1),  C_OR),
            ("BACKGROUND",    (1,0),(1,-1),  C_NAVY),
            ("BACKGROUND",    (2,0),(2,-1),  HexColor("#2A4A6A")),
            ("INNERGRID",     (0,0),(-1,-1), 2.0, C_WH),
        ]))
        story.append(_kc)
        story.append(Spacer(1, 0.35*cm))

        if not _cd.empty:
            _cost_hdr = [["Machine","Date","Shift","Key Failure",
                           "Spare Part / Ref","Qty","Unit Price (€)","Total Cost (€)"]]
            _cost_rows_pdf = []
            _cost_extras   = []
            for _ri, (_idx, _row) in enumerate(_cd.iterrows()):
                _tc = float(pd.to_numeric(_row.get("Total Part Cost",0), errors="coerce") or 0)
                _cost_rows_pdf.append([
                    str(_row.get(COL_MACHINE, "—"))[:12],
                    str(_row.get(COL_DATE, "—"))[:10],
                    str(_row.get("Shift", "—"))[:4],
                    str(_row.get("Key Failure","—"))[:28],
                    str(_row.get("Spare Part Ref","—"))[:22],
                    str(int(pd.to_numeric(_row.get("Qty",0), errors="coerce") or 0)),
                    f"€ {float(pd.to_numeric(_row.get('Unit Price (€)',0),errors='coerce') or 0):.2f}",
                    f"€ {_tc:.2f}",
                ])
                _ri_t = len(_cost_rows_pdf)
                if _tc > 0:
                    _cost_extras += [
                        ("TEXTCOLOR", (7,_ri_t),(7,_ri_t), C_OR),
                        ("FONTNAME",  (7,_ri_t),(7,_ri_t), "Helvetica-Bold"),
                    ]
            _cws_cost = [IW*x for x in [0.10,0.10,0.06,0.24,0.20,0.06,0.12,0.12]]
            story.append(_tbl(_cost_hdr+_cost_rows_pdf, _cws_cost, fs=7, extras=_cost_extras))
            story.append(Spacer(1, 0.3*cm))

            _tot_t = Table([[
                Paragraph("TOTAL SPARE PARTS EXPENDITURE",
                          ps("totp", fontSize=11, fontName="Helvetica-Bold", textColor=C_WH)),
                Paragraph(f"€ {_total_spend:,.2f}",
                          ps("totv", fontSize=14, fontName="Helvetica-Bold",
                             textColor=C_WH, alignment=2)),
            ]], colWidths=[IW * 0.65, IW * 0.35])
            _tot_t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), C_OR),
                ("TOPPADDING",    (0,0),(-1,-1), 12),
                ("BOTTOMPADDING", (0,0),(-1,-1), 12),
                ("LEFTPADDING",   (0,0),(-1,-1), 16),
                ("RIGHTPADDING",  (0,0),(-1,-1), 16),
                ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ]))
            story.append(_tot_t)
        else:
            story.append(Paragraph("No spare parts costs recorded for this period.", S_BODY))

        def add_footer(canvas_obj, doc):
            canvas_obj.saveState()
            canvas_obj.setFont("Helvetica", 7)
            canvas_obj.setFillColorRGB(0.63, 0.47, 0.34)
            canvas_obj.drawCentredString(
                W/2, 1.35*cm,
                f"≡ TE CONNECTIVITY  ·  STAMPING CMMS  ·  TANGIER     |     "
                f"Mean MTTR · Mean MTBF · Availability = MTBF/(MTBF+MTTR)     |     "
                f"{datetime.now().strftime('%d/%m/%Y')}     |     "
                f"Page {doc.page}")
            canvas_obj.setStrokeColorRGB(0.91, 0.40, 0.04)
            canvas_obj.setLineWidth(1.2)
            canvas_obj.line(MARGIN, 1.85*cm, W-MARGIN, 1.85*cm)
            canvas_obj.restoreState()

        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=1.4*cm, bottomMargin=2.5*cm)
        doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
        return buf.getvalue()

    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        try:
            _pdf_bytes = build_pdf(df, kpi, ma_disp, pareto)
            st.download_button(
                "DOWNLOAD PDF REPORT", data=_pdf_bytes,
                file_name=f"TE_CMMS_{today_str}.pdf", mime="application/pdf",
                use_container_width=True)
        except Exception as _e:
            st.warning(f"PDF: `pip install reportlab` ({_e})")
    with ec2:
        try:
            _xl = export_excel(df, kpi)
            st.download_button(
                "EXCEL MULTI-SHEET", data=_xl,
                file_name=f"TE_CMMS_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as _e:
            st.warning(f"Excel: {_e}")
    with ec3:
        st.download_button(
            "CSV PARETO",
            data=pareto.to_csv(index=False, sep=";").encode("utf-8"),
            file_name=f"TE_pareto_{today_str}.csv",
            mime="text/csv", use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 — STOPS QUALIFICATION
# ══════════════════════════════════════════════════════════════════════════════
with tab_qual:

    def _is_qualified(r):
        return any(str(r.get(c, "")).strip() not in ("", "None", "nan")
                   for c in ["Shift", "Key Failure",
                              "Issue Description", "Action Taken", "Spare Part Ref"])

    _check_cols = [c for c in ["Shift","Key Failure"] if c in df.columns]
    _qual_n = int(df[_check_cols].apply(_is_qualified, axis=1).sum()) if _check_cols else 0
    _stop_n = int((df["mttr_h"] > 0).sum())
    _pct_q  = (_qual_n / _stop_n * 100) if _stop_n > 0 else 0

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{TE_BLACK} 0%,#2A1A0A 100%);
                border:2px solid {TE_ORANGE};border-radius:14px;
                padding:20px 28px;margin-bottom:18px">
      <div style="display:flex;align-items:center;justify-content:space-between;
                  flex-wrap:wrap;gap:16px">
        <div>
          <div style="font-family:'Barlow Condensed',sans-serif;font-size:20px;font-weight:800;
                      color:{TE_ORANGE};letter-spacing:2px;text-transform:uppercase;margin-bottom:6px">
             Stops Qualification
          </div>
          <div style="font-family:'JetBrains Mono',monospace;font-size:10px;
                      color:rgba(255,255,255,0.6);line-height:1.9">
            Fill in <strong style="color:{TE_ORANGE2}">User ID</strong>,
            <strong style="color:{TE_ORANGE2}">Shift</strong>
            and <strong style="color:{TE_ORANGE2}">Key Failure</strong>
            · Click <strong style="color:{TE_ORANGE2}"> Save Changes</strong> to persist on disk
          </div>
        </div>
        <div style="text-align:right">
          <div style="font-family:'Barlow Condensed',sans-serif;font-size:34px;font-weight:800;
                      color:{TE_ORANGE};line-height:1">
            {_qual_n}<span style="font-size:16px;color:rgba(255,255,255,0.35)">/{_stop_n}</span>
          </div>
          <div style="font-size:9px;color:rgba(255,255,255,0.4);
                      font-family:'JetBrains Mono',monospace;letter-spacing:1px">
            STOPS QUALIFIED
          </div>
        </div>
      </div>
      <div style="margin-top:14px">
        <div style="display:flex;justify-content:space-between;font-size:9px;
                    color:rgba(255,255,255,0.4);font-family:'JetBrains Mono',monospace;
                    margin-bottom:5px">
          <span>Qualification progress</span>
          <span style="color:{TE_ORANGE}">{_qual_n} / {_stop_n} ({_pct_q:.0f}%)</span>
        </div>
        <div style="background:rgba(255,255,255,0.1);border-radius:4px;height:7px;overflow:hidden">
          <div style="background:linear-gradient(90deg,{TE_ORANGE},{TE_DARK});height:100%;
                      width:{min(_pct_q,100):.1f}%;border-radius:4px;
                      transition:width 0.4s ease"></div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if os.path.exists(PERSISTENT_CSV):
        _fsize = os.path.getsize(PERSISTENT_CSV)
        _fmod  = datetime.fromtimestamp(os.path.getmtime(PERSISTENT_CSV))
        st.markdown(f"""
        <div style="background:#eafaf1;border:1px solid #a9dfbf;
                    border-left:4px solid {TE_GREEN};border-radius:8px;
                    padding:8px 14px;margin-bottom:10px;font-size:11px;
                    color:#145a32;display:flex;align-items:center;gap:10px">
            <span>Persistent file active: <strong>{PERSISTENT_CSV}</strong>
          &nbsp;·&nbsp; {_fsize/1024:.1f} KB
          &nbsp;·&nbsp; Last saved: <strong>{_fmod.strftime('%d/%m/%Y %H:%M')}</strong></span>
        </div>
        """, unsafe_allow_html=True)

    # ── Quick Search ──
    st.markdown(f"""
    <div style="background:{TE_WHITE};border:1px solid #EDE0D4;
                border-left:4px solid {TE_ORANGE};border-radius:10px;
                padding:14px 18px;margin-bottom:14px">
      <div style="font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:700;
                  letter-spacing:2.5px;text-transform:uppercase;
                  color:{TE_ORANGE};margin-bottom:10px">
         Quick Search
      </div>
    """, unsafe_allow_html=True)

    _df_stops = df[df["mttr_h"] > 0].copy()
    _fcol1, _fcol2 = st.columns(2)

    with _fcol1:
        _machines_avail = ["All"] + sorted(_df_stops[COL_MACHINE].dropna().unique().tolist())
        _filter_machine = st.selectbox(" Machine ID", options=_machines_avail,
                                       index=0, key="q_filter_machine")
    with _fcol2:
        _dates_raw    = pd.to_datetime(_df_stops[COL_DATE], errors="coerce").dropna()
        _dates_avail  = sorted(_dates_raw.dt.date.unique())
        _date_options = ["All"] + [d.strftime("%m/%d/%Y") for d in _dates_avail]
        _filter_date_str = st.selectbox(" Exact Date", options=_date_options,
                                         index=0, key="q_filter_date")

    st.markdown("</div>", unsafe_allow_html=True)

    _display_cols = [c for c in [
        COL_MACHINE, COL_DATE, COL_STATUS, "mttr_h",
        "User ID", "Shift", "Key Failure",
        "Issue Description", "Action Taken",
        "Spare Part Ref", "Qty", "Unit Price (€)", "Total Part Cost",
    ] if c in df.columns]

    _df_base = _df_stops.copy()
    if _filter_machine != "All":
        _df_base = _df_base[_df_base[COL_MACHINE] == _filter_machine]
    if _filter_date_str != "All":
        _target_date = pd.to_datetime(_filter_date_str, format="%m/%d/%Y", errors="coerce")
        if pd.notna(_target_date):
            _df_base = _df_base[
                pd.to_datetime(_df_base[COL_DATE], errors="coerce").dt.date
                == _target_date.date()]

    _orig_idx = _df_base.index.values
    _df_show  = _df_base[_display_cols].copy()
    if COL_DATE in _df_show.columns:
        _df_show[COL_DATE] = (
            pd.to_datetime(_df_show[COL_DATE], errors="coerce")
            .dt.strftime("%m/%d/%Y").fillna("—"))
    _df_show = _df_show.reset_index(drop=True)

    _n_shown     = len(_df_show)
    _is_filtered = _filter_machine != "All" or _filter_date_str != "All"
    st.markdown(
        f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:9px;'
        f'color:#9A7A60;margin-bottom:8px;letter-spacing:1px">'
        f'Showing: <strong style="color:{TE_ORANGE}">{_n_shown}</strong>'
        f' stop{"s" if _n_shown != 1 else ""}'
        f'{"  ·  active filter / " + str(_stop_n) + " total stops" if _is_filtered else "  ·  " + str(_stop_n) + " total stops"}'
        f'</div>',
        unsafe_allow_html=True)

    if _df_show.empty:
        st.info("No stops match the selected filters.")
    else:
        if "Qty" in _df_show.columns and "Unit Price (€)" in _df_show.columns:
            _df_show["Total Part Cost"] = (
                pd.to_numeric(_df_show["Qty"], errors="coerce").fillna(0) *
                pd.to_numeric(_df_show["Unit Price (€)"], errors="coerce").fillna(0.0)
            ).round(2)

        _edited = st.data_editor(
            _df_show,
            use_container_width=True,
            height=min(700, max(250, _n_shown * 42 + 62)),
            num_rows="fixed",
            column_order=_display_cols,
            column_config={
                COL_MACHINE: st.column_config.TextColumn(
                    " Machine", disabled=True, width="small"),
                COL_DATE: st.column_config.TextColumn(
                    " Date", disabled=True, width="small"),
                COL_STATUS: st.column_config.TextColumn(
                    " Status", disabled=True, width="medium"),
                "mttr_h": st.column_config.NumberColumn(
                    " MTTR (h)", format="%.4f", disabled=True, width="small"),
                "User ID": st.column_config.TextColumn(
                    " User ID", disabled=False, width="small",
                    max_chars=20, help="Your technician badge / employee ID"),
                "Shift": st.column_config.SelectboxColumn(
                    " Shift", options=SHIFTS, required=False, width="small",
                    help="A (6-14h) · B (14-22h) · C (22-6h)"),
                "Key Failure": st.column_config.SelectboxColumn(
                    " Key Failure", options=KEY_FAILURES,
                    required=False, width="large", help="Root cause of the stop"),
                "Issue Description": st.column_config.TextColumn(
                    " Issue Description", disabled=False, width="large", max_chars=300),
                "Action Taken": st.column_config.TextColumn(
                    " Action Taken", disabled=False, width="large", max_chars=300),
                "Spare Part Ref": st.column_config.TextColumn(
                    " Spare Part / Ref", disabled=False, width="medium", max_chars=100),
                "Qty": st.column_config.NumberColumn(
                    " Qty", disabled=False, width="small",
                    min_value=0, step=1, default=0),
                "Unit Price (€)": st.column_config.NumberColumn(
                    " Unit Price (€)", disabled=False, width="small",
                    min_value=0.0, step=0.01, format="%.2f", default=0.0),
                "Total Part Cost": st.column_config.NumberColumn(
                    " Total Cost (€)", disabled=True, width="small",
                    format="%.2f", help="Qty × Unit Price (auto-calculated on Save)"),
            },
            key="qual_editor_v6"
        )

        if _edited is not None and "Qty" in _edited.columns and "Unit Price (€)" in _edited.columns:
            _live_cost = float(
                (pd.to_numeric(_edited["Qty"], errors="coerce").fillna(0) *
                 pd.to_numeric(_edited["Unit Price (€)"], errors="coerce").fillna(0.0)
                ).sum())
            if _live_cost > 0:
                st.markdown(f"""
                <div style="background:linear-gradient(90deg,#0d1f0d,#1a3a1a);
                            border:1.5px solid #27AE60;border-radius:8px;
                            padding:10px 18px;margin:6px 0;
                            display:flex;align-items:center;gap:12px">
                  <span style="font-size:20px"></span>
                  <div>
                    <span style="font-family:'Barlow Condensed',sans-serif;
                                 font-size:10px;color:rgba(255,255,255,0.5);
                                 letter-spacing:2px;text-transform:uppercase">
                      Current view · Spare Parts Cost
                    </span><br>
                    <span style="font-family:'Barlow Condensed',sans-serif;
                                 font-size:22px;font-weight:800;color:#27AE60">
                      € {_live_cost:,.2f}
                    </span>
                    <span style="font-size:10px;color:rgba(255,255,255,0.4);margin-left:8px">
                      (click  Save to update global KPI and persist to disk)
                    </span>
                  </div>
                </div>""", unsafe_allow_html=True)

        # ── Save button ──
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        _sb_left, _sb_mid, _sb_right = st.columns([1.5, 2, 1.5])
        with _sb_mid:
            _save_clicked = st.button(
                "  SAVE CHANGES", type="primary",
                use_container_width=True, key="btn_save_qual",
                help="Save entries and write to disk (tpm_data_persistent.csv)")

        # ── Excel export ──
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        def _build_excel_export():
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            _src = st.session_state.edited_df.copy()
            _INTERNAL = {"mttr_h", "mtbf_h", "date_only"}
            for _c in ["User ID", "Shift", "Key Failure",
                       "Issue Description", "Action Taken", "Spare Part Ref"]:
                if _c in _src.columns:
                    _src[_c] = (_src[_c].fillna("").astype(str).str.strip()
                                        .replace({"nan": "", "None": ""}))
            for _c in ["Qty", "Unit Price (€)", "Total Part Cost"]:
                if _c in _src.columns:
                    _src[_c] = pd.to_numeric(_src[_c], errors="coerce").fillna(0)
            if "Qty" in _src.columns and "Unit Price (€)" in _src.columns:
                _src["Total Part Cost"] = (_src["Qty"] * _src["Unit Price (€)"]).round(2)
            if "mttr_h" in _src.columns:
                _src["MTTR (h)"] = _src["mttr_h"].round(4)
            if "mtbf_h" in _src.columns:
                _src["MTBF (h)"] = _src["mtbf_h"].round(4)
            if COL_DATE in _src.columns:
                _parsed = pd.to_datetime(_src[COL_DATE], errors="coerce")
                _src[COL_DATE] = _parsed.dt.strftime("%m/%d/%Y").fillna("")
            _hydra_base = [c for c in _src.columns
                           if c not in _INTERNAL
                           and c not in ["User ID","Shift","Key Failure",
                                         "Issue Description","Action Taken",
                                         "Spare Part Ref","Qty",
                                         "Unit Price (€)","Total Part Cost",
                                         "MTTR (h)","MTBF (h)"]]
            _metric_cols = [c for c in ["MTTR (h)","MTBF (h)"] if c in _src.columns]
            _qual_cols   = [c for c in [
                "User ID","Shift","Key Failure","Issue Description","Action Taken",
                "Spare Part Ref","Qty","Unit Price (€)","Total Part Cost",
            ] if c in _src.columns]
            _all_cols = _hydra_base + _metric_cols + _qual_cols

            def _style_sheet(ws, df_data, header_color="1B2A4A", zebra_color="F7F4F0"):
                _hdr_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                _hdr_fill   = PatternFill("solid", fgColor=header_color)
                _hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                _zebra_fill = PatternFill("solid", fgColor=zebra_color)
                _thin       = Side(style="thin", color="D0C0B0")
                _border     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                for cell in ws[1]:
                    cell.font = _hdr_font; cell.fill = _hdr_fill
                    cell.alignment = _hdr_align; cell.border = _border
                ws.row_dimensions[1].height = 30
                for row_idx, row in enumerate(ws.iter_rows(
                        min_row=2, max_row=ws.max_row), start=2):
                    _fill = (_zebra_fill if row_idx % 2 == 0 else PatternFill())
                    for cell in row:
                        cell.fill = _fill; cell.border = _border
                        cell.alignment = Alignment(vertical="center")
                for col_cells in ws.columns:
                    _w = max((len(str(c.value or "")) for c in col_cells), default=8)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(_w+3, 45)
                ws.freeze_panes = "A2"

            _buf = io.BytesIO()
            with pd.ExcelWriter(_buf, engine="openpyxl",
                                 datetime_format="MM/DD/YYYY") as _writer:
                _df_full = _src[_all_cols].copy()
                _df_full.to_excel(_writer, sheet_name="Full Dataset", index=False)
                _style_sheet(_writer.sheets["Full Dataset"], _df_full)

                # ── Stops sheet: ONLY real stops (mttr_h > 0), NO production rows ──
                _mask_stops = (
                    pd.to_numeric(_src.get("mttr_h", pd.Series([0]*len(_src))),
                                  errors="coerce").fillna(0) > 0
                )
                _stops = _src[_mask_stops].copy().reset_index(drop=True)
                _stops_cols = [c for c in [
                    COL_MACHINE, COL_DATE, COL_STATUS, "MTTR (h)",
                    "User ID", "Shift", "Key Failure",
                    "Issue Description", "Action Taken",
                    "Spare Part Ref", "Qty", "Unit Price (€)", "Total Part Cost",
                ] if c in _stops.columns]

                def _mark_qual(r):
                    return (str(r.get("Shift","")).strip() not in ("","nan","None") or
                            str(r.get("Key Failure","")).strip() not in ("","nan","None"))
                _stops["Qualified"] = _stops[_stops_cols].apply(
                    _mark_qual, axis=1).map({True:" Yes", False:" No"})
                _stops2_cols = _stops_cols + ["Qualified"]
                _stops[_stops2_cols].to_excel(_writer, sheet_name="Stops", index=False)
                _style_sheet(_writer.sheets["Stops"], _stops[_stops2_cols],
                             header_color="E8650A")

                _cost_rows = _stops[
                    pd.to_numeric(_stops.get("Total Part Cost", pd.Series([0])),
                                  errors="coerce").fillna(0) > 0
                ].copy() if "Total Part Cost" in _stops.columns else pd.DataFrame()
                if not _cost_rows.empty:
                    _cost_cols = [c for c in [
                        COL_MACHINE, COL_DATE, "User ID", "Key Failure",
                        "Spare Part Ref", "Qty", "Unit Price (€)", "Total Part Cost",
                    ] if c in _cost_rows.columns]
                    _cost_df = _cost_rows[_cost_cols].copy()
                    _total_row = {c: "" for c in _cost_cols}
                    _total_row["Spare Part Ref"] = "TOTAL"
                    _total_row["Total Part Cost"] = _cost_df["Total Part Cost"].sum().round(2)
                    _cost_df = pd.concat([_cost_df, pd.DataFrame([_total_row])], ignore_index=True)
                    _cost_df.to_excel(_writer, sheet_name="Cost Summary", index=False)
                    _ws_cost = _writer.sheets["Cost Summary"]
                    _style_sheet(_ws_cost, _cost_df)
                    _last = _ws_cost.max_row
                    for _cell in _ws_cost[_last]:
                        _cell.fill = PatternFill("solid", fgColor="E8650A")
                        _cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            _buf.seek(0)
            return _buf.read()

        _ex_left, _ex_mid, _ex_right = st.columns([1.5, 2, 1.5])
        with _ex_mid:
            try:
                _excel_bytes = _build_excel_export()
                _ts_xl = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    label="EXPORT QUALIFIED DATA (EXCEL)",
                    data=_excel_bytes,
                    file_name=f"TE_Qualification_{_ts_xl}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_export_excel"
                )
            except Exception as _ex_err:
                st.warning(f" Excel export unavailable: {_ex_err}")

        # ── SAVE LOGIC ──
        if _save_clicked and _edited is not None:
            _n_saved = 0
            _ALL_SAVE_COLS = [
                "User ID", "Shift", "Key Failure",
                "Issue Description", "Action Taken", "Spare Part Ref",
                "Qty", "Unit Price (€)",
            ]
            for _col in _ALL_SAVE_COLS:
                if _col not in _edited.columns:
                    continue
                _new_vals = _edited[_col].values
                _old_vals = st.session_state.edited_df.loc[_orig_idx, _col].values
                try:
                    _diff = ~(
                        (pd.isnull(_new_vals) & pd.isnull(_old_vals)) |
                        (np.array(_new_vals, dtype=object) == np.array(_old_vals, dtype=object))
                    )
                    for _k, _orig_i in enumerate(_orig_idx):
                        if _diff[_k]:
                            st.session_state.edited_df.at[_orig_i, _col] = _new_vals[_k]
                            _n_saved += 1
                except Exception:
                    st.session_state.edited_df.loc[_orig_idx, _col] = _new_vals
                    _n_saved += len(_orig_idx)

            _edf2 = st.session_state.edited_df
            _edf2["Total Part Cost"] = (
                pd.to_numeric(_edf2["Qty"], errors="coerce").fillna(0) *
                pd.to_numeric(_edf2["Unit Price (€)"], errors="coerce").fillna(0.0)
            ).round(2)
            st.session_state.edited_df = _edf2
            save_persistent(st.session_state.edited_df)
            st.session_state["_save_result"] = _n_saved
            st.rerun()

        if st.session_state.get("_save_result") is not None:
            _nr = st.session_state.pop("_save_result")
            if _nr > 0:
                st.markdown(f"""
                <div style="background:#eafaf1;border:1.5px solid #a9dfbf;
                            border-left:5px solid {TE_GREEN};border-radius:10px;
                            padding:14px 20px;margin-top:6px;
                            display:flex;align-items:center;gap:14px">
                  <span style="font-size:24px"></span>
                  <div>
                    <div style="font-family:'Barlow Condensed',sans-serif;font-size:15px;
                                font-weight:800;color:#1e8449;text-transform:uppercase;
                                letter-spacing:1px;margin-bottom:3px">
                      Changes saved — written to disk
                    </div>
                    <div style="font-size:12px;color:#145a32;line-height:1.6">
                      <strong>{_nr}</strong> field{"s" if _nr != 1 else ""}
                      modified · Saved to <code>{PERSISTENT_CSV}</code>
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="background:#fff8e1;border:1.5px solid #ffe082;
                            border-left:5px solid {TE_AMBER};border-radius:10px;
                            padding:12px 18px;margin-top:6px;
                            display:flex;align-items:center;gap:10px">
                  <span style="font-size:18px"></span>
                  <span style="font-size:12px;color:#5d4037">
                    No changes detected compared to last save.
                  </span>
                </div>
                """, unsafe_allow_html=True)

    # ── Qualified Stops Summary ──
    if _qual_n > 0:
        st.markdown(f'<div class="te-section">Qualified Stops Summary</div>',
                    unsafe_allow_html=True)
        _q_all = df[df[["Shift","Key Failure"]].apply(_is_qualified, axis=1)]
        _ts    = datetime.now().strftime("%Y%m%d_%H%M")
        _cost_total_q = 0.0
        if st.session_state.edited_df is not None and "Total Part Cost" in st.session_state.edited_df.columns:
            _cost_total_q = float(
                pd.to_numeric(st.session_state.edited_df["Total Part Cost"],
                              errors="coerce").fillna(0).sum())

        _qm1, _qm2, _qm3, _qm4 = st.columns(4)
        with _qm1:
            st.metric(" Qualified Stops", f"{_qual_n} / {_stop_n}",
                      delta=f"{_pct_q:.0f}% of total")
        with _qm2:
            _uid_n = int((_q_all.get("User ID", pd.Series(dtype=str))
                           .astype(str).str.strip()
                           .replace({"":pd.NA,"nan":pd.NA,"None":pd.NA}).notna()).sum())
            st.metric(" With User ID", f"{_uid_n} rows",
                      delta="signed" if _uid_n > 0 else "none")
        with _qm3:
            _kf_n = int((_q_all.get("Key Failure", pd.Series(dtype=str))
                          .astype(str).str.strip()
                          .replace({"":pd.NA,"nan":pd.NA,"None":pd.NA}).notna()).sum())
            st.metric(" Key Failure filled", f"{_kf_n} rows")
        with _qm4:
            st.metric(" Spare Parts Cost", f"€ {_cost_total_q:,.2f}",
                      delta="recorded" if _cost_total_q > 0 else "no costs yet")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        _q_exp_cols = [c for c in [
            COL_MACHINE, COL_DATE, COL_STATUS, "mttr_h", "User ID", "Shift", "Key Failure",
            "Issue Description", "Action Taken",
            "Spare Part Ref", "Qty", "Unit Price (€)", "Total Part Cost",
        ] if c in _q_all.columns]

        _ec1, _ec2 = st.columns(2)
        with _ec1:
            st.download_button(
                f"CSV  {_qual_n} QUALIFIED STOP(S)",
                data=_q_all[_q_exp_cols].to_csv(index=False, sep=";").encode("utf-8"),
                file_name=f"TE_qualified_stops_{_ts}.csv",
                mime="text/csv", use_container_width=True)
        with _ec2:
            st.info(f" **{_qual_n}** qualified stop(s) · "
                    f"Full PDF report available in ** KPIs** tab · "
                    f"Data persisted to **{PERSISTENT_CSV}**")


# ──────────────────────────────────────────────────────────────────────────────
#  FOOTER
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="text-align:center;font-family:'JetBrains Mono',monospace;font-size:9px;
            letter-spacing:2px;color:#C0A080;padding:24px 0 12px;
            border-top:1px solid #E0D0C0;margin-top:32px">
    ≡ TE CONNECTIVITY · STAMPING CMMS · TANGIER<br>
    Mean MTTR = Total Downtime / Failures &nbsp;·&nbsp;
    Mean MTBF = Total Uptime / Failures &nbsp;·&nbsp;
    Availability = MTBF / (MTBF + MTTR)<br>
    {datetime.now().strftime('%m/%d/%Y')}
</div>
""", unsafe_allow_html=True)
